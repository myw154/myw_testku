# coding: utf-8

import sys

import openpyxl


class GenExcel():
    def __init__(self, row_l):
        self.row_l = row_l
        self.file = openpyxl.Workbook()
        # self.file.remove_sheet(self.file.active)

    def process(self, tab_file):
        for line in open(tab_file):
            line_l = line.strip('\n').split('\t')
            table_name = line_l[0] + line_l[1]
            worksheets = [i.title for i in self.file.worksheets]
            if not table_name:
                table_name = '无学段学科'
            if table_name in worksheets:
                table = self.file[table_name]
            else:
                table = self.file.create_sheet()
                table.title = table_name
                for index, row in enumerate(self.row_l):
                    table.cell(1, index+1, row)
            max_row = table.max_row
            for index, info in enumerate(line_l):
                table.cell(max_row+1, index+1, info)
        self.file.save('%s.xlsx' % tab_file)


def remind():
    if len(sys.argv) < 3:
        print('参数错误:')
        print('示例如下:')
        print('\tpython %s tab_file row1 row2 row3' % sys.argv[0])
        sys.exit(-1)


def main():
    remind()
    row_l = sys.argv[2:]
    gen_excel = GenExcel(row_l)
    gen_excel.process(sys.argv[1])


if __name__ == '__main__':
    main()
