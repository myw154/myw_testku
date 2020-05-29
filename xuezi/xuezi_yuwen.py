import requests
import re
import os
import json
from retrying import retry
from lxml import etree
from requests import request
from openpyxl import load_workbook


class BookListYuWen(object):

    def __init__(self):
        self.url_list = []
        self.book_url = []
        self.book_first_url = []
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36'
        }
        self.headers_dzkb = {
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3',
            'Accept-Encoding': 'gzip, deflate',
            'Accept-Language': 'zh-CN,zh;q=0.9',
            'Connection': 'keep-alive',
            'Host': 'www.dzkbw.com',
            'Upgrade-Insecure-Requests': '1',
            'Cookie': 'ASPSESSIONIDAQSDDTCR=PNODCHKBCOCBGOFPAHEFELJN; _ga=GA1.2.1041419005.1569411852; _gid=GA1.2.1303393327.1569411852; Hm_lvt_463b13a51a26c92e0b49cd9e0dedb52b=1569411854; ASPSESSIONIDASSDCTDQ=LGDBIFLBPJEDOCCFNDDKEPEG; ASPSESSIONIDCSSADBCD=MPAAKLMBIGBFEBCLPNOJJFEM; ASPSESSIONIDCQQCDTCR=BCFFPGACFFKPBEOAABPIIAEN; ASPSESSIONIDCSRBARBS=FCIJEJACMOENGAABODEJJAEC; ASPSESSIONIDCQSBCACD=OAAAOIACBHBMFKPMDNJMOFGP; ASPSESSIONIDCSSADACD=HBGGNJBCGLJFNCKPFJCAHLHP; ASPSESSIONIDCSSACADC=KELCFACCHDIFGLOODAIMBPLP; ASPSESSIONIDASRDCSCR=CGEJILCCAJKCMBCMDMMMKFDD; ASPSESSIONIDASSBDADD=CJBGEDDCOJFGDENNNIBNNIBI; ASPSESSIONIDCSRCDSCQ=IBNOHEDCBCFCCAOEHPIKMGJN; ASPSESSIONIDCQRCBACD=NDOFENDCFJFCJNGMFANONOLD; ASPSESSIONIDCSTAAQBS=NFKAHODCHNABKLECPJKNLMGF; ASPSESSIONIDAQQBAQAT=IDNALKECICKPIABCCHCJOAFA; ASPSESSIONIDASRCDTDR=KPNCEMFCDOJMDKJPNAPKJKHB; ASPSESSIONIDAQSBCBDD=PAGBBBGCIMEHGBBCECHLHFKJ; ASPSESSIONIDASSDBTCR=FLJMCIGCHLJFEDAHJIIMLCOA; ASPSESSIONIDASRBARAT=IKKAJAHCFPEBHGKFKCECFLMH; ASPSESSIONIDCQSCCBCD=HJFFMDHCKPLNJHCCIIDLOLLD; ASPSESSIONIDCSSDCACD=MIDDBCLCPBENHPBOOOAJKMPK; ASPSESSIONIDASSBBQAT=JEDPFJLCIGNPAPJAIKFFNMHC; ASPSESSIONIDAQTDCBCC=LHIIDLLCEGNOJBPPKPPAMECH; ASPSESSIONIDASRCCTCQ=PANMOLLCPLJJGMFCEDANBIDD; ASPSESSIONIDAQTBADDD=ACKKLKMCDEEDFDPKCJIALFMA; ASPSESSIONIDASRDCBDC=NJIKLFNCDLEKDMCNNBHJPFON; ASPSESSIONIDCSTCBBCD=KLMDOMNCNPNAMGKGPPJECOFL; _gat_gtag_UA_66838525_1=1; Hm_lpvt_463b13a51a26c92e0b49cd9e0dedb52b=1569569659',
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.75 Safari/537.36'
        }
        # /Users/guo/Desktop/mywtest/xuezi_net
        self.current_path_name = os.path.dirname(os.path.abspath(__file__))

    def create_file_path(self):
        if not os.path.exists(self.current_path_name+'/yuwen'):
            os.makedirs(self.current_path_name+'/yuwen')
        if not os.path.exists(self.current_path_name+'/yuwen/img'):
            os.makedirs(self.current_path_name+'/yuwen/img')
        if not os.path.exists(self.current_path_name+'/yuwen/chapter_html'):
            os.makedirs(self.current_path_name+'/yuwen/chapter_html')

    @retry(stop_max_attempt_number=6, stop_max_delay=3000, wait_fixed=3000, wait_incrementing_increment=500)
    def send_request(self, url, headers=None):
        requests.packages.urllib3.disable_warnings()
        res = request(method='get', url=url, headers=headers, timeout=20, verify=False)
        return res.content

    def struct_url(self):
        open_excel = load_workbook(os.path.abspath('book_list.xlsx'))
        tab =open_excel['book_list']
        max_num = tab.max_row
        # for i in range():
        for i in range(1, max_num+1):
            url_1 = tab.cell(i, 1).value
            url_2 = tab.cell(i, 2).value
            url_3 = tab.cell(i, 3).value
            url_4 = tab.cell(i, 4).value
            url_re_obj = re.search(r'http://www.yuwenziyuan.com/.*isstop=True', url_4)
            if url_re_obj:
                self.book_url.append({'url_1': url_1, 'url_2': url_2, 'url_3': url_3, 'url_4': url_4})

    # 把字节转成字符串
    def bytes_to_html(self, html_bytes):
        try:
            html_obj = etree.HTML(html_bytes.decode(encoding='utf-8'))
        except Exception:
            try:
                html_obj = etree.HTML(html_bytes.decode(encoding='gbk'))
            except Exception:
                try:
                    html_obj = etree.HTML(html_bytes.decode(encoding='GB2312'))
                except Exception:
                    html_obj = etree.HTML(html_bytes.decode(encoding='GB18030'))
        return html_obj

    # 转换成html字符串
    def bytes_to_str(self, html_bytes):
        try:
            html_obj = html_bytes.decode(encoding='utf-8')
        except Exception:
            try:
                html_obj = html_bytes.decode(encoding='gbk')
            except Exception:
                try:
                    html_obj = html_bytes.decode(encoding='GB2312')
                except Exception:
                    html_obj = html_bytes.decode(encoding='GB18030')
        return html_obj

    def anay_subj(self, html_bytes):
        html_obj = self.bytes_to_html(html_bytes)
        info_list = html_obj.xpath('.//div[@class="daohan"]/*/text()')
        # 获得原始书本的章节目录
        mulu_list = []
        mulu_a_list = html_obj.xpath('.//div[@class="bookmulu"]/a')
        for a_obj in mulu_a_list:
            name = a_obj.xpath('./*/text() | ./text()')[0] if len(a_obj.xpath('./*/text() | ./text()')) > 0 else None
            href = 'http://www.dzkbw.com'+a_obj.xpath('./@href')[0]
            mulu_list.append([name, href])
        return info_list, mulu_list

    # 构造每一本书的请求地址
    def struct_book_url(self):
        # 构造请求的地址
        self.create_file_path()
        self.struct_url()

    # 获得一个章节的新的地址
    def get_new_url(self, url):
        res_bytes = self.send_request(url, self.headers_dzkb)
        html_obj = self.bytes_to_html(res_bytes)
        # 获得二维码的地址
        if html_obj.xpath('.//div[@class="bookimg"]//a[2]/@href'):
            if not html_obj.xpath('.//div[@class="bookimg"]//a[2]/@href')[0].startswith('/go.asp?'):
                seract_url = 'http://www.dzkbw.com'+html_obj.xpath('.//div[@class="bookimg"]//a[1]/@href')[0]
            else:
                seract_url = 'http://www.dzkbw.com' + html_obj.xpath('.//div[@class="bookimg"]//a[2]/@href')[0]
        else:
            seract_url = 'http://www.dzkbw.com' + html_obj.xpath('.//div[@class="bookimg"]//a[1]/@href')[0]
        seract_res_bytes = self.send_request(seract_url, self.headers_dzkb)
        seract_obj_html = self.bytes_to_html(seract_res_bytes)
        new_chapter_url = seract_obj_html.xpath('.//head/meta/@content')[1].split("'")[1]
        return new_chapter_url

    def get_all_chapter(self, book_url, page_num):
        # 获得每个章节的新地址
        pages_all = {}
        if book_url in ['http://www.dzkbw.com/books/sjb/yuwen/9s/015.htm','http://www.dzkbw.com/books/sjb/yuwen/9x/027.htm','http://www.dzkbw.com/books/sjb/yuwen/7x/013.htm']:
            start_num = 9
        elif book_url == 'http://www.dzkbw.com/books/sjb/yuwen/8s/017.htm':
            start_num = 14
        elif book_url == 'http://www.dzkbw.com/books/sjb/yuwen/8x/014.htm':
            start_num = 10
        else:
            start_num = 1
        for i in range(start_num, page_num + 1):
            if i <= 9:
                chapter_url = re.sub(r'/\d+.htm', '/00' + str(i) + '.htm', book_url)
            elif 10 <= i <= 99:
                chapter_url = re.sub(r'/\d+.htm', '/0' + str(i) + '.htm', book_url)
            else:
                chapter_url = re.sub(r'/\d+.htm', '/' + str(i) + '.htm', book_url)
            print(chapter_url)
            new_url = self.get_new_url(chapter_url)
            res_obj = self.send_request(new_url)
            html_str = self.bytes_to_str(res_obj)
            html_obj = self.bytes_to_html(res_obj)
            html_list = [html_str]
            a_url_list = set(html_obj.xpath('.//div[@class="ckqw"]/div[@style="text-align:center"]/a/@href'))
            if a_url_list:
                for a_url in a_url_list:
                    son_url = '/'.join(new_url.split('/')[:-1])+'/'+a_url
                    print(son_url)
                    res_son_obj = self.send_request(son_url)
                    son_html = self.bytes_to_str(res_son_obj)
                    html_list.append({son_url: son_html})
            pages_all[chapter_url] = {new_url: html_list}
        return pages_all


    # 获得每本书的目录信息，获得每本书的第一页的url地址
    def get_book_list(self):
        self.struct_book_url()
        print(len(self.book_url))
        for url_info_dict in self.book_url:
            # 获得原始目录的html
            try:
                star_url = url_info_dict['url_2'].replace('\n', '')
                print(star_url)
                html_bytes = self.send_request(url=star_url, headers=self.headers_dzkb)
                html_str = self.bytes_to_str(html_bytes)
                url_info_dict['chapter_html'] = html_str
                page_num = int(re.search(r'maxPage=(\d+);', html_str).group(1))
                url_info_dict['pages_all_num'] = page_num
                print(page_num)
                pages_all = self.get_all_chapter(star_url, page_num)
                url_info_dict['pages'] = pages_all
                with open(self.current_path_name+'/yuwen/already.txt', 'a+', encoding='utf-8') as f:
                    f.write(json.dumps(url_info_dict, ensure_ascii=False)+'\n')
            except Exception:
                with open(self.current_path_name+'/yuwen/url_error.txt', 'a+', encoding='utf-8') as f:
                    f.write(json.dumps(url_info_dict)+'\n')

    def get_error(self):
        with open(self.current_path_name+'/yuwen/url_error.txt', 'r', encoding='utf-8') as f:
            #f.write(json.dumps(url_info_dict)+'\n')
            url_info_dict_list = f.readlines()
            print(len(url_info_dict_list))
            for url_dict_str in url_info_dict_list:
                url_info_dict = eval(url_dict_str)
                try:
                    star_url = url_info_dict['url_2'].replace('\n', '')
                    print(star_url)
                    html_bytes = self.send_request(url=star_url, headers=self.headers_dzkb)
                    html_str = self.bytes_to_str(html_bytes)
                    url_info_dict['chapter_html'] = html_str
                    page_num = int(re.search(r'maxPage=(\d+);', html_str).group(1))
                    url_info_dict['pages_all_num'] = page_num
                    print(page_num)
                    pages_all = self.get_all_chapter(star_url, page_num)
                    url_info_dict['pages'] = pages_all
                    with open(self.current_path_name+'/yuwen/already.txt', 'a+', encoding='utf-8') as f:
                        f.write(json.dumps(url_info_dict, ensure_ascii=False)+'\n')
                except Exception:
                    with open(self.current_path_name+'/yuwen/url_error2.txt', 'a+', encoding='utf-8') as f:
                        f.write(json.dumps(url_info_dict)+'\n')


if __name__ == '__main__':
    yuwen = BookListYuWen()
    #yuwen.get_book_list()
    yuwen.get_error()

