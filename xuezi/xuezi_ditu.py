import requests
import re
import os
import json
from retrying import retry
from lxml import etree
from requests import request
from openpyxl import load_workbook


class BookListDiTu(object):

    def __init__(self):
        self.url_list = []
        self.book_url = []
        self.book_first_url = []
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36'
        }
        self.headers_dzkb = {
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3',
            'Accept-Encoding': 'gzip, deflate',
            'Accept-Language': 'zh-CN,zh;q=0.9',
            'Connection': 'keep-alive',
            'Host': 'www.dzkbw.com',
            'Upgrade-Insecure-Requests': '1',
            'Cookie': 'ASPSESSIONIDAQSDDTCR=PNODCHKBCOCBGOFPAHEFELJN; _ga=GA1.2.1041419005.1569411852; _gid=GA1.2.1303393327.1569411852; Hm_lvt_463b13a51a26c92e0b49cd9e0dedb52b=1569411854; ASPSESSIONIDASSDCTDQ=LGDBIFLBPJEDOCCFNDDKEPEG; ASPSESSIONIDCSSADBCD=MPAAKLMBIGBFEBCLPNOJJFEM; ASPSESSIONIDCQQCDTCR=BCFFPGACFFKPBEOAABPIIAEN; ASPSESSIONIDCSRBARBS=FCIJEJACMOENGAABODEJJAEC; ASPSESSIONIDCQSBCACD=OAAAOIACBHBMFKPMDNJMOFGP; ASPSESSIONIDCSSADACD=HBGGNJBCGLJFNCKPFJCAHLHP; ASPSESSIONIDCSSACADC=KELCFACCHDIFGLOODAIMBPLP; ASPSESSIONIDASRDCSCR=CGEJILCCAJKCMBCMDMMMKFDD; ASPSESSIONIDASSBDADD=CJBGEDDCOJFGDENNNIBNNIBI; ASPSESSIONIDCSRCDSCQ=IBNOHEDCBCFCCAOEHPIKMGJN; ASPSESSIONIDCQRCBACD=NDOFENDCFJFCJNGMFANONOLD; ASPSESSIONIDCSTAAQBS=NFKAHODCHNABKLECPJKNLMGF; ASPSESSIONIDAQQBAQAT=IDNALKECICKPIABCCHCJOAFA; ASPSESSIONIDASRCDTDR=KPNCEMFCDOJMDKJPNAPKJKHB; ASPSESSIONIDAQSBCBDD=PAGBBBGCIMEHGBBCECHLHFKJ; ASPSESSIONIDASSDBTCR=FLJMCIGCHLJFEDAHJIIMLCOA; ASPSESSIONIDASRBARAT=IKKAJAHCFPEBHGKFKCECFLMH; ASPSESSIONIDCQSCCBCD=HJFFMDHCKPLNJHCCIIDLOLLD; ASPSESSIONIDCSSDCACD=MIDDBCLCPBENHPBOOOAJKMPK; ASPSESSIONIDASSBBQAT=JEDPFJLCIGNPAPJAIKFFNMHC; ASPSESSIONIDAQTDCBCC=LHIIDLLCEGNOJBPPKPPAMECH; ASPSESSIONIDASRCCTCQ=PANMOLLCPLJJGMFCEDANBIDD; ASPSESSIONIDAQTBADDD=ACKKLKMCDEEDFDPKCJIALFMA; ASPSESSIONIDASRDCBDC=NJIKLFNCDLEKDMCNNBHJPFON; ASPSESSIONIDCSTCBBCD=KLMDOMNCNPNAMGKGPPJECOFL; _gat_gtag_UA_66838525_1=1; Hm_lpvt_463b13a51a26c92e0b49cd9e0dedb52b=1569569659',
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.75 Safari/537.36'
        }
        # /Users/guo/Desktop/mywtest/xuezi_net
        self.current_path_name = os.path.dirname(os.path.abspath(__file__))

    def create_file_path(self):
        if not os.path.exists(self.current_path_name+'/ditu'):
            os.makedirs(self.current_path_name+'/ditu')
        if not os.path.exists(self.current_path_name+'/ditu/img'):
            os.makedirs(self.current_path_name+'/ditu/img')
        if not os.path.exists(self.current_path_name+'/ditu/book_chapter_html'):
            os.makedirs(self.current_path_name+'/ditu/book_chapter_html')
        if not os.path.exists(self.current_path_name + '/ditu/original_chapter_html/'):
            os.makedirs(self.current_path_name + '/ditu/original_chapter_html/')

    @retry(stop_max_attempt_number=5, stop_max_delay=3000, wait_fixed=3000, wait_incrementing_increment=500)
    def send_request(self, url, headers=None):
        requests.packages.urllib3.disable_warnings()
        res = request(method='get', url=url, headers=headers, timeout=10, verify=False)
        return res.content

    def struct_url(self):
        open_excel = load_workbook(os.path.abspath('book_list.xlsx'))
        tab = open_excel['book_list']
        max_num = tab.max_row
        for i in range(1, max_num):
        #for i in range(1, max_num+1):
            url_1 = tab.cell(i, 1).value
            url_2 = tab.cell(i, 2).value
            url_3 = tab.cell(i, 3).value
            url_4 = tab.cell(i, 4).value
            re_obj = re.search(r'http://www.ditu.cn/(.+)isstop=True|http://www.ywcbs.com/(.*)', url_4)
            if re_obj:
                self.book_url.append({'url_1': url_1, 'url_2': url_2, 'url_3': url_3, 'url_4': url_4})

    # 把字节转成字符串
    def bytes_to_html(self, html_bytes):
        try:
            html_obj = etree.HTML(html_bytes.decode(encoding='utf-8'))
        except Exception:
            try:
                html_obj = etree.HTML(html_bytes.decode(encoding='gbk'))
            except Exception:
                try:
                    html_obj = etree.HTML(html_bytes.decode(encoding='GB2312'))
                except Exception:
                    html_obj = etree.HTML(html_bytes.decode(encoding='GB18030'))
        return html_obj

    def bytes_to_str(self, html_bytes):
        try:
            html_obj = html_bytes.decode(encoding='utf-8')
        except Exception:
            try:
                html_obj = html_bytes.decode(encoding='gbk')
            except Exception:
                try:
                    html_obj = html_bytes.decode(encoding='GB2312')
                except Exception:
                    html_obj = html_bytes.decode(encoding='GB18030')
        return html_obj

    def anay_subj(self, html_bytes):
        html_obj = self.bytes_to_html(html_bytes)
        info_list = html_obj.xpath('.//div[@class="daohan"]/*/text()')
        # 获得原始书本的章节目录
        mulu_list = []
        mulu_a_list = html_obj.xpath('.//div[@class="bookmulu"]/a')
        for a_obj in mulu_a_list:
            name = a_obj.xpath('./*/text() | ./text()')[0] if len(a_obj.xpath('./*/text() | ./text()')) > 0 else None
            href = 'http://www.dzkbw.com'+a_obj.xpath('./@href')[0]
            mulu_list.append([name, href])
        return info_list, mulu_list

    # 构造每一本书的请求地址
    def struct_book_url(self):
        # 创建一些必要的文件夹路径
        self.create_file_path()
        # 构造书本的地址
        self.struct_url()

    def get_pdf_book(self, book_url):
        res_bye = self.send_request(book_url, self.headers)
        html_str = res_bye.decode(encoding='utf-8')
        pdf_url = 'http://www.ditu.cn'+re.search(r'href="(/uploads/.*pdf)"', html_str).group(1)
        return pdf_url

    def get_jpg_book(self, book_url):
        res_bytes = self.send_request(book_url)
        html_str = res_bytes.decode(encoding='utf-8')
        page_num = int(re.search('bookConfig.totalPageCount=(\d+);', html_str).group(1))
        return page_num


    def download_img(self, book_url, page_num):
        book_name = '_'.join(book_url.split('/')[4:-1])
        if not os.path.exists(self.current_path_name+'/ditu/img/'+book_name):
            os.makedirs(self.current_path_name+'/ditu/img/'+book_name)
        for i in range(1, page_num+1):
            img_url = '/'.join(book_url.split('/')[:-1]) + '/files/mobile/' + str(i) + '.jpg'
            try:
                img_bytes = self.send_request(url=img_url)
                img_name = '_'.join(img_url.split('/')[-4:])
                with open(self.current_path_name+'/ditu/img/'+book_name+'/'+img_name, 'wb') as f:
                    f.write(img_bytes)
            except Exception:
                with open(self.current_path_name+'/ditu/img_url_error.txt', 'a', encoding='utf-8') as f:
                    f.write(str([book_url, img_url])+'\n')


    # 获得每本书的目录信息，获得每本书的第一页的url地址
    def get_book_list(self):
        self.struct_book_url()
        print(len(self.book_url))
        for url_info_dict in self.book_url:
            # 获得原始目录的html
            try:
                star_url = url_info_dict['url_1'].replace('\n', '')
                html_bytes = self.send_request(url=star_url, headers=self.headers_dzkb)
                chapter_html = self.bytes_to_str(html_bytes)
                url_info_dict['chapter_html'] = chapter_html

                star_url_2 = url_info_dict['url_2'].replace('\n', '')
                html_bytes_2 = self.send_request(url=star_url_2, headers=self.headers_dzkb)
                chapter_html_2 = self.bytes_to_str(html_bytes_2)
                url_info_dict['chapter_html_2'] = chapter_html_2

                book_url = url_info_dict['url_4'].split("'")[1]
                print(book_url)
                if 'http://www.ditu.cn' in book_url:
                    pdf_url = self.get_pdf_book(book_url)
                    url_info_dict['pdf_url'] = pdf_url
                if 'http://www.ywcbs.com/ywcbsFile' in book_url:
                    # 获得书本的页码数
                    new_book_url = '/'.join(book_url.split('/')[:-1])+'/mobile/index.html'
                    page_num = self.get_jpg_book(new_book_url)
                    url_info_dict['page_all_num'] = page_num
                    url_info_dict['img_url'] = '/'.join(book_url.split('/')[:-1])+'/files/mobile/num.jpg num范围[1:page_all_num]'
                    # 下载图片
                # 保存已经获得的信息
                with open(self.current_path_name+'/ditu/already_get_book.txt', 'a+', encoding='utf-8') as f:
                    f.write(json.dumps(url_info_dict, ensure_ascii=False)+'\n')
            except Exception:
                with open(self.current_path_name+'/ditu/error_url.txt', 'a+', encoding='utf-8') as f:
                    f.write(json.dumps(url_info_dict)+'\n')



    def analys_ditu(self):
        with open(self.current_path_name + '/ditu/already_get_book.txt', 'r', encoding='utf-8') as f:
            info_dcit_str_list = f.readlines()
            i = 1
            for info_dcit_str in info_dcit_str_list:
                info_dict = eval(info_dcit_str)
                book_info = {}
                book_info['source_url'] = info_dict['url_2']
                chapter_html = info_dict['chapter_html_2']
                html_obj = etree.HTML(chapter_html)
                base_info = html_obj.xpath('.//div[@class="daohan"]//a/text()')
                book_info['book_name'] = '_'.join(base_info)
                if re.search(r'一年级|二年级|三年级|四年级|五年级|六年级', base_info[3]):
                    book_info['period'] = '小学'
                    book_info['grade'] = re.search(r'一年级|二年级|三年级|四年级|五年级|六年级', base_info[3]).group()
                elif re.search(r'七年级|八年级|九年级', base_info[3]):
                    book_info['period'] = '初中'
                    book_info['grade'] = re.search('七年级|八年级|九年级', base_info[3]).group()
                elif re.search(r'高一|高二|高三', base_info[3]):
                    book_info['period'] = '高中'
                    book_info['grade'] = re.search(r'高一|高二|高三', base_info[3]).group()
                book_info['press_version'] = base_info[1]
                book_info['subject'] = base_info[2]
                book_info['catalogs'] = []
                book_info['pages'] = []
                if 'http://www.ditu.cn' in info_dict['url_4'].split("'")[1]:
                    chapter_html = info_dict['chapter_html_2']
                    html_obj = etree.HTML(chapter_html)
                    a_obj_list = html_obj.xpath('.//div[@class="mululist"]/a')
                    chapter_list = []
                    for a_obj in a_obj_list:
                        name_list = a_obj.xpath('./text()|./*/text()')
                        if len(name_list) > 1:
                            name = ' '.join(name_list)
                        else:
                            name = name_list[0]
                        chapter_list.append(name)
                    big_chapter_list = []
                    n = len(chapter_list)
                    big_index_list = []
                    for chapter_name, i in zip(chapter_list, range(1, n + 1)):
                        if re.search('封面|目录|第一章|第二章|第三章|第四章|第五章|第六章|第七章|第八章|第九章|第十章|附录', chapter_name[:5]):
                            big_chapter_list.append({'name': chapter_name, 'catalogs': [], 'pages': []})
                            big_index_list.append(i)
                    big_index_list.append(n + 1)

                    # 把小章节添加到大章节中
                    m = len(big_index_list)
                    for pre_big_dcit, j in zip(big_chapter_list, range(0, m)):
                        for little_chaper in chapter_list[big_index_list[j]:big_index_list[j + 1] - 1]:
                            pre_big_dcit['catalogs'].append({'name': little_chaper, 'catalogs': [], 'pages': []})
                    book_info['catalogs'] = big_chapter_list
                    # img_path 是 /home/mengyanwei/myw_env/ditu/pdf_to_img+pdf名字
                    pdf_name = '_'.join(info_dict['pdf_url'].split('/')[3:]).replace('.pdf', '')
                    img_list = os.listdir(os.path.join('/home/mengyanwei/myw_env/ditu/pdf_to_img/'+pdf_name))
                    img_list.sort(key=lambda x: int(re.search(r'_(\d+)\.jpg', x).group(1)))
                    book_info['pages'].extend(img_list)
                    print(json.dumps(book_info, ensure_ascii=False))
                    with open(self.current_path_name + '/ditu/add_analys_already_book.json', 'a', encoding='utf-8') as f_json:
                        f_json.write(json.dumps(book_info, ensure_ascii=False)+'\n')
                if 'http://www.ywcbs.com/ywcbsFile' in info_dict['url_4'].split("'")[1]:
                    # 提取列表
                    chapter_html = info_dict['chapter_html_2']
                    html_obj = etree.HTML(chapter_html)
                    a_obj_list = html_obj.xpath('.//div[@class="mululist"]/a')
                    chapter_list = []
                    chapter_name_list = []
                    for a_obj in a_obj_list:
                        name_list = a_obj.xpath('./text()|./*/text()')
                        href = a_obj.xpath('./@href')[0]
                        if len(name_list) > 1:
                            name = ' '.join(name_list)
                        else:
                            name = name_list[0]
                        chapter_list.append([name, href])
                        chapter_name_list.append(name)

                    # 识别大章节
                    big_chapter_list = []
                    n = len(chapter_list)
                    big_index_list = []
                    if re.search('第[一二三四五六七八九十]单元', ' '.join(chapter_name_list)):
                        for chapter_name, i in zip(chapter_list, range(1, n + 1)):
                            if re.search('封面|目录|[一二三四五六七八九十]+单元', chapter_name[0][:5]):
                                big_chapter_list.append({'name': chapter_name[0], 'catalogs': [], 'pages': int(re.search('/(\d+)\.htm', chapter_name[1]).group(1))-1})
                                big_index_list.append(i)
                        big_index_list.append(n + 1)

                        # 把小章节添加到大章节中
                        m = len(big_index_list)
                        for pre_big_dcit, j in zip(big_chapter_list, range(0, m)):
                            for little_chaper in chapter_list[big_index_list[j]:big_index_list[j + 1] - 1]:
                                pre_big_dcit['catalogs'].append({'name': little_chaper, 'catalogs': [], 'pages': []})
                        book_info['catalogs'] = big_chapter_list

                    # 如果没有提取到大章节
                    else:
                        for name in chapter_list:
                            book_info['catalogs'].append({'name': name[0], 'catalogs': [], 'pages': int(re.search('/(\d+)\.htm', name[1]).group(1))-1})

                    pages_all_num = info_dict['page_all_num']
                    # print(info_dict['img_url'])
                    img_pre_url = info_dict['img_url'].replace(' num范围[1:page_all_num]', '')
                    for i in range(1, pages_all_num+1):
                        book_info['pages'].append(img_pre_url.replace('num', str(i)))
                    print(json.dumps(book_info, ensure_ascii=False))
                    with open(self.current_path_name + '/ditu/add_analys_already_book.json', 'a', encoding='utf-8') as f_json:
                        f_json.write(json.dumps(book_info, ensure_ascii=False)+'\n')
                    print(i)
                    i += 1
if __name__ == '__main__':
    ditu = BookListDiTu()
    #ditu.get_book_list()
    ditu.analys_ditu()
