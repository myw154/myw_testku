import requests
import re
import os
import json
from retrying import retry
from lxml import etree
from requests import request
from openpyxl import load_workbook


class BookList171English(object):

    def __init__(self):
        self.url_list = []
        self.book_url = []
        self.book_first_url = []
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36'
        }
        self.headers_dzkb = {
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3',
            'Accept-Encoding': 'gzip, deflate',
            'Accept-Language': 'zh-CN,zh;q=0.9',
            'Connection': 'keep-alive',
            'Host': 'www.dzkbw.com',
            'Upgrade-Insecure-Requests': '1',
            'Cookie': 'ASPSESSIONIDAQSDDTCR=PNODCHKBCOCBGOFPAHEFELJN; _ga=GA1.2.1041419005.1569411852; _gid=GA1.2.1303393327.1569411852; Hm_lvt_463b13a51a26c92e0b49cd9e0dedb52b=1569411854; ASPSESSIONIDASSDCTDQ=LGDBIFLBPJEDOCCFNDDKEPEG; ASPSESSIONIDCSSADBCD=MPAAKLMBIGBFEBCLPNOJJFEM; ASPSESSIONIDCQQCDTCR=BCFFPGACFFKPBEOAABPIIAEN; ASPSESSIONIDCSRBARBS=FCIJEJACMOENGAABODEJJAEC; ASPSESSIONIDCQSBCACD=OAAAOIACBHBMFKPMDNJMOFGP; ASPSESSIONIDCSSADACD=HBGGNJBCGLJFNCKPFJCAHLHP; ASPSESSIONIDCSSACADC=KELCFACCHDIFGLOODAIMBPLP; ASPSESSIONIDASRDCSCR=CGEJILCCAJKCMBCMDMMMKFDD; ASPSESSIONIDASSBDADD=CJBGEDDCOJFGDENNNIBNNIBI; ASPSESSIONIDCSRCDSCQ=IBNOHEDCBCFCCAOEHPIKMGJN; ASPSESSIONIDCQRCBACD=NDOFENDCFJFCJNGMFANONOLD; ASPSESSIONIDCSTAAQBS=NFKAHODCHNABKLECPJKNLMGF; ASPSESSIONIDAQQBAQAT=IDNALKECICKPIABCCHCJOAFA; ASPSESSIONIDASRCDTDR=KPNCEMFCDOJMDKJPNAPKJKHB; ASPSESSIONIDAQSBCBDD=PAGBBBGCIMEHGBBCECHLHFKJ; ASPSESSIONIDASSDBTCR=FLJMCIGCHLJFEDAHJIIMLCOA; ASPSESSIONIDASRBARAT=IKKAJAHCFPEBHGKFKCECFLMH; ASPSESSIONIDCQSCCBCD=HJFFMDHCKPLNJHCCIIDLOLLD; ASPSESSIONIDCSSDCACD=MIDDBCLCPBENHPBOOOAJKMPK; ASPSESSIONIDASSBBQAT=JEDPFJLCIGNPAPJAIKFFNMHC; ASPSESSIONIDAQTDCBCC=LHIIDLLCEGNOJBPPKPPAMECH; ASPSESSIONIDASRCCTCQ=PANMOLLCPLJJGMFCEDANBIDD; ASPSESSIONIDAQTBADDD=ACKKLKMCDEEDFDPKCJIALFMA; ASPSESSIONIDASRDCBDC=NJIKLFNCDLEKDMCNNBHJPFON; ASPSESSIONIDCSTCBBCD=KLMDOMNCNPNAMGKGPPJECOFL; _gat_gtag_UA_66838525_1=1; Hm_lpvt_463b13a51a26c92e0b49cd9e0dedb52b=1569569659',
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.75 Safari/537.36'
        }
        # /Users/guo/Desktop/mywtest/xuezi_net
        self.proxy_dict = {
            'http': 'http://H4069413V6797J9D:FC20F1BDEAF3A3BC@http-dyn.abuyun.com:9020'
        }
        self.current_path_name = os.path.dirname(os.path.abspath(__file__))

    def create_file_path(self):
        if not os.path.exists(self.current_path_name+'/english_171'):
            os.makedirs(self.current_path_name+'/english_171')
        if not os.path.exists(self.current_path_name+'/english_171/img'):
            os.makedirs(self.current_path_name+'/english_171/img')
        if not os.path.exists(self.current_path_name+'/english_171/chapter_html'):
            os.makedirs(self.current_path_name+'/english_171/chapter_html')
        if not os.path.exists(self.current_path_name + '/english_171/original_chapter_html/'):
            os.makedirs(self.current_path_name + '/english_171/original_chapter_html/')

    @retry(stop_max_attempt_number=5, stop_max_delay=3000, wait_fixed=3000, wait_incrementing_increment=500)
    def send_request(self, url, headers=None):
        requests.packages.urllib3.disable_warnings()
        res = request(method='get', url=url, headers=headers, timeout=10, verify=False,)
        return res

    def struct_url(self):
        open_excel = load_workbook(os.path.abspath('book_list.xlsx'))
        tab =open_excel['book_list']
        max_num = tab.max_row
       # for i in range(801):
        for i in range(1, max_num+1):
            url_1 = tab.cell(i, 1).value
            url_2 = tab.cell(i, 2).value
            url_3 = tab.cell(i, 3).value
            url_4 = tab.cell(i, 4).value
            url_re_obj = re.search('http://www\.171english\.cn/(.+)\.html\?isstop=True', url_4)
            if url_re_obj:
                self.book_url.append({'url_1': url_1, 'url_2': url_2, 'url_3': url_3, 'url_4': url_4})

    # 把字节转成字符串
    def bytes_to_html(self, html_bytes):
        try:
            html_obj = etree.HTML(html_bytes.content.decode(encoding='utf-8'))
        except Exception:
            try:
                html_obj = etree.HTML(html_bytes.content.decode(encoding='gbk'))
            except Exception:
                try:
                    html_obj = etree.HTML(html_bytes.content.decode(encoding='GB2312'))
                except Exception:
                    html_obj = etree.HTML(html_bytes.content.decode(encoding='GB18030'))
        return html_obj

    # 转换成html字符串
    def bytes_to_str(self, html_bytes):
        try:
            html_obj = html_bytes.content.decode(encoding='utf-8')
        except Exception:
            try:
                html_obj = html_bytes.content.decode(encoding='gbk')
            except Exception:
                try:
                    html_obj = html_bytes.content.decode(encoding='GB2312')
                except Exception:
                    html_obj = html_bytes.content.decode(encoding='GB18030')
        return html_obj

    def anay_subj(self, html_bytes):
        html_obj = self.bytes_to_html(html_bytes)
        info_list = html_obj.xpath('.//div[@class="daohan"]/*/text()')
        # 获得原始书本的章节目录
        mulu_list = []
        mulu_a_list = html_obj.xpath('.//div[@class="bookmulu"]/a')
        for a_obj in mulu_a_list:
            name = a_obj.xpath('./*/text() | ./text()')[0] if len(a_obj.xpath('./*/text() | ./text()')) > 0 else None
            href = 'http://www.dzkbw.com'+a_obj.xpath('./@href')[0]
            mulu_list.append([name, href])
        return info_list, mulu_list

    # 构造每一本书的请求地址
    def struct_book_url(self):
        # 构造请求的地址
        self.create_file_path()
        self.struct_url()

    # 下载图片
    def download_img(self, book_url, chapter_url, img_url, sound_url, son_chpater_info):
        book_name = '_'.join(book_url.split('/')[3:-1])
        if not os.path.exists(self.current_path_name+'/english_171/img/'+book_name):
            os.makedirs(self.current_path_name+'/english_171/img/'+book_name)
        img_name = '_'.join(img_url.split('/')[3:])
        try:
            img_path = self.current_path_name+'/english_171/img/'+book_name+'/'+img_name
            res_bytes = self.send_request(img_url)
            with open(img_path, 'wb') as f:
                f.write(res_bytes)
            with open(self.current_path_name + '/english_171/book_chapter_img_url.txt', 'a', encoding='utf-8') as f:
                f.write(str([book_url, chapter_url, img_url, sound_url, son_chpater_info])+'\n')
        except Exception:
            with open(self.current_path_name+'/english_171/img_url_error.txt', 'a', encoding='utf-8') as f:
                f.write(str([book_url, chapter_url, img_url, sound_url, son_chpater_info])+'\n')

    # 获得书本的各个章节
    def get_book(self, book_url, chapter_url):
        try:
            res_bytes = self.send_request(chapter_url, self.headers)
            html_obj = self.bytes_to_html(res_bytes)
            # 获得音频的地址
            sound_list = html_obj.xpath('.//source/@src')
            if sound_list:
                sound_url = '/'.join(chapter_url.split('/')[:-1]) + '/' + sound_list[0]
            else:
                sound_url = 'none'
            # 获得该章节的图片地址
            img_url_list = html_obj.xpath('.//div[@id="zoom"]//img/@src')
            for img_url_short in img_url_list:
                img_url = '/'.join(chapter_url.split('/')[:-1])+'/'+img_url_short
                self.download_img(book_url, chapter_url, img_url, sound_url)

            # 保存每个章节的html页面
            book_name = '_'.join(book_url.split('/')[3:-1])
            if not os.path.exists(self.current_path_name + '/english_171/chapter_html/' + book_name):
                os.makedirs(self.current_path_name + '/english_171/chapter_html/' + book_name)
            chapter_html_name = re.search(r'http://www.171english.cn/(.+).html', chapter_url).group(1).replace('/', '_')
            with open(self.current_path_name+'/english_171/chapter_html/'+book_name+'/'+chapter_html_name+'.txt', 'wb') as f:
                f.write(res_bytes)

            # 当进入到第六页的时候会进入死循环，所以修改为lesson7
            if chapter_url == 'http://www.171english.cn/fltrp/7A/book/lesson6.html':
                next_chapter_url = 'http://www.171english.cn/fltrp/7A/book/lesson7.html'
                self.get_book(book_url, next_chapter_url)
            if chapter_url == 'http://www.171english.cn/fltrp/9A/revisiona.html':
                next_chapter_url = 'http://www.171english.cn/fltrp/9A/lesson7.html'
                self.get_book(book_url, next_chapter_url)
            if chapter_url == 'http://www.171english.cn/fltrp/9b/lesson6.html':
                next_chapter_url = 'http://www.171english.cn/fltrp/9b/lesson7.html'
                self.get_book(book_url, next_chapter_url)
            else:
                # 获得下一章节的地址
                if chapter_url != 'http://www.171english.cn/fltrp/9b/lesson6.html': 
                    a_content = html_obj.xpath('.//div[@class="context"]//li[2]//a/text()')
                    a_url_list = html_obj.xpath('.//div[@class="context"]//li[2]//a/@href')
                    if a_url_list and (a_content[0] != '返回目录' and a_content[0] != '课本封面和目录' and a_content[0] != '语音和不规则动词'):
                        next_chapter_url = re.sub(r'/less(.+).html\?isstop=True|/unit(.+).html\?isstop=True|/sless(.+).html\?isstop=True', '/'+a_url_list[0], book_url)
                        self.get_book(book_url, next_chapter_url)
        except Exception:
            with open(self.current_path_name+'/english_171/chapter_url_error.txt', 'a', encoding='utf-8') as f:
                f.write(str([book_url, chapter_url])+'\n')

    # 获得含有nes书本的目录信息
    def get_new_mulu(self, book_url):
        res_bytes = self.send_request(book_url, self.headers_dzkb)
        html_obj = self.bytes_to_html(res_bytes)
        new_mulu_list = []
        # 网页会更新 有时候需要[2:]
        a_obj_list = html_obj.xpath('.//div[@class="mululist"]//a')[1:]
        for a_obj in a_obj_list:
            a_href = 'http://www.dzkbw.com'+a_obj.xpath('./@href')[0]
            name = a_obj.xpath('./text() | ./*/text()')[0]
            new_mulu_list.append([a_href, name])
        return new_mulu_list


    # 请求获得新的地址，拿到章节内容
    def get_new_content(self, book_url, chapter_url):
        book_name = '_'.join(book_url.split('/')[3:-1])
        res_bytes = self.send_request(chapter_url)
        html_obj = self.bytes_to_html(res_bytes)
        # 获得音频的地址
        sound_list = html_obj.xpath('.//source/@src')
        if sound_list:
            sound_url = '/'.join(chapter_url.split('/')[:-1])+'/'+sound_list[0]
        else:
            sound_url = 'none'
        content = html_obj.xpath('.//div[@id="zoom"]')[0]
        content_str = etree.tostring(content, encoding='utf-8').decode(encoding='utf-8')
        if not os.path.exists(self.current_path_name+'/english_171/img/'+book_name):
            os.makedirs(self.current_path_name+'/english_171/img/'+book_name)
        chapter_name = '_'.join(chapter_url.split('/')[3:]).replace('html?isstop=True', 'txt')
        with open(self.current_path_name+'/english_171/img/'+book_name+'/'+chapter_name, 'w', encoding='utf-8') as f:
            f.write(content_str)
        with open(self.current_path_name + '/english_171/book_chapter_nes_url.txt', 'a', encoding='utf-8') as f:
            f.write(str([book_url, chapter_url, sound_url]) + '\n')

    # 获得一含有nes 的书本信息
    def get_new_book(self, book_url):
        new_mulu_list = self.get_new_mulu(book_url)
        for a_href_list in new_mulu_list:
            try:
                new_url = self.get_new_url(a_href_list[0])
                self.get_new_content(book_url, new_url)
            except Exception:
                with open(self.current_path_name+'/english_171/chapter_nes_url_error.txt', 'a', encoding='utf-8') as f:
                    f.write(str([book_url, a_href_list[0]])+'\n')
        return new_mulu_list

    # 获得每本书的目录信息，获得每本书的第一页的url地址
    def get_book_list(self):
        self.struct_book_url()
        print(len(self.book_url))

        for url_info_dict in self.book_url:
            # 获得原始目录的html
            try:
                star_url = url_info_dict['url_1'].replace('\n', '')
                html_bytes = self.send_request(url=star_url, headers=self.headers_dzkb)
                url_info_dict['chapter_html'] = self.bytes_to_str(html_bytes)

                book_url = re.search(r'http://www.171english.cn/(.+).html\?isstop=True', url_info_dict['url_4']).group()
                if 'http://www.171english.cn/nse' in book_url:
                    new_mulu_list = self.get_new_book(url_info_dict['url_2'])
                    url_info_dict['new_chapter'] = new_mulu_list
                print(book_url)
                if 'xuanxiu' in book_url:
                    chapter_url = re.sub(r'/less(.+).html|/unit(.+).html', '/lesson1.html', book_url)
                if 'http://www.171english.cn/shangdong' in book_url:
                    chapter_url = re.sub(r'/less(.+).html|/unit(.+).html', '/lesson1.html', book_url)
                else:
                    chapter_url = re.sub(r'/less(.+)\.html|/unit(.+).html|/sless(.+)\.html', '/lesson.html', book_url)
                self.get_book(book_url, chapter_url)
                with open(self.current_path_name+'/english_171/already_get_book.txt', 'a+', encoding='utf-8') as f:
                    f.write(json.dumps(url_info_dict)+'\n')

                # 原始网站的页面html保存
                html_name = '_'.join(star_url.split('/')[3:-1])+'.txt'
                with open(self.current_path_name+'/english_171/original_chapter_html/'+html_name, 'wb') as f:
                    f.write(html_bytes)
            except Exception:
                with open(self.current_path_name+'/english_171/error_url.txt', 'a+', encoding='utf-8') as f:
                    f.write(json.dumps(url_info_dict)+'\n')

    # 获得一个章节的新的地址
    def get_new_url(self, url):
        res_bytes = self.send_request(url, self.headers_dzkb)
        html_obj = self.bytes_to_html(res_bytes)
        # 获得二维码的地址
        if html_obj.xpath('.//div[@class="bookimg"]//a[2]/@href'):
            if not html_obj.xpath('.//div[@class="bookimg"]//a[2]/@href')[0].startswith('/go.asp?'):
                seract_url = 'http://www.dzkbw.com'+html_obj.xpath('.//div[@class="bookimg"]//a[1]/@href')[0]
            else:
                seract_url = 'http://www.dzkbw.com' + html_obj.xpath('.//div[@class="bookimg"]//a[2]/@href')[0]
        else:
            seract_url = 'http://www.dzkbw.com' + html_obj.xpath('.//div[@class="bookimg"]//a[1]/@href')[0]
        seract_res_bytes = self.send_request(seract_url, self.headers_dzkb)
        seract_obj_html = self.bytes_to_html(seract_res_bytes)
        new_chapter_url = seract_obj_html.xpath('.//head/meta/@content')[1].split("'")[1]
        return new_chapter_url

    def get_all_pages(self, book_url, pages_all_num):
        pages_all_list = []
        for i in range(1, pages_all_num+1):
            if i <= 9:
                chapter_url = re.sub(r'/\d+.htm', '/00{}.htm'.format(i), book_url)
            elif 10 <= i <= 99:
                chapter_url = re.sub(r'/\d+.htm', '/0{}.htm'.format(i), book_url)
            else:
                chapter_url = re.sub(r'/\d+.htm', '/{}.htm'.format(i), book_url)
            print(chapter_url)
            new_chapter_url = self.get_new_url(chapter_url)

            res = self.send_request(new_chapter_url)
            if res.status_code != 404:
                pages_all_list.append({chapter_url: self.bytes_to_str(res)})
        return pages_all_list

    # 获得每本书的目录信息，获得每本书的第一页的url地址
    def get_book_list2(self):
        self.struct_book_url()
        print(len(self.book_url))
        for url_info_dict in self.book_url:
            # 获得原始目录的html
            try:
                book_url = url_info_dict['url_2']
                print('*'*30)
                print(book_url)
                html_bytes = self.send_request(url=book_url, headers=self.headers_dzkb)
                html_str = self.bytes_to_str(html_bytes)
                url_info_dict['chapter_html'] = html_str
                pages_all_num = int(re.search(r'maxPage=(\d+);', html_str).group(1))
                print(pages_all_num)
                url_info_dict['pages_all_num'] = pages_all_num
                pages_all_list = self.get_all_pages(book_url, pages_all_num)
                url_info_dict['pages'] = pages_all_list
                with open(self.current_path_name+'/english_171/already_get_book.txt', 'a+', encoding='utf-8') as f:
                    f.write(json.dumps(url_info_dict, ensure_ascii=False)+'\n')
            except Exception:
                with open(self.current_path_name+'/english_171/error_url.txt', 'a+', encoding='utf-8') as f:
                    f.write(json.dumps(url_info_dict, ensure_ascii=False)+'\n')

    def analys_info(self):
        with open(self.current_path_name + '/english_171/already_get_book2.txt', 'r', encoding='utf-8') as f:
            # f.write(json.dumps(url_info_dict, ensure_ascii=False) + '\n')
            info_dict_str_list = f.readlines()
        i = 0
        for info_dict_str in info_dict_str_list:
            info_dict = eval(info_dict_str)
            try:
                print(info_dict['url_2'])
                book_info = {}
                book_info['source_url']=info_dict['url_2']
                chapter_html = info_dict['chapter_html']
                html_obj = etree.HTML(chapter_html)
                base_info = html_obj.xpath('.//div[@class="daohan"]//a/text()')
                book_info['book_info'] = '_'.join(base_info)
                if re.search(r'一年级|二年级|三年级|四年级|五年级|六年级', base_info[3]):
                    book_info['period'] = '小学'
                    book_info['grade'] = re.search(r'一年级|二年级|三年级|四年级|五年级|六年级', base_info[3]).group()
                elif re.search(r'七年级|八年级|九年级', base_info[3]):
                    book_info['period'] = '初中'
                    book_info['grade'] = re.search('七年级|八年级|九年级', base_info[3]).group()
                elif re.search(r'高一|高二|高三', base_info[3]):
                    book_info['period'] = '高中'
                    book_info['grade'] = re.search(r'高一|高二|高三', base_info[3]).group()
                book_info['press_version'] = base_info[1]
                book_info['subject'] = base_info[2]
                book_info['catalogs'] = []
                book_info['pages'] = []
                # print(book_info)
                # 提取章节列表
                a_obj_list = html_obj.xpath('.//div[@class="mululist"]//a')[1:]
                for a_obj in a_obj_list:
                    name_list = a_obj.xpath('./text()|./*/text()')
                    if len(name_list) > 1:
                        name = ' '.join(name_list)
                    else:
                        name = name_list[0]
                    a_href = 'http://www.dzkbw.com'+a_obj.xpath('./@href')[0]
                    book_info['catalogs'].append({'name': name, 'catalogs': [], 'pages': [], 'a_href': a_href})
                # print(book_info)

                all_pages_list = info_dict['pages']
                # 把原始存储的列表字典，改成pages的字典，便于获得每个章节的html
                new_all_pages_dict = {}
                for page_dict in all_pages_list:
                    for key, value in page_dict.items():
                        new_all_pages_dict[key] = value

                # 遍历每个章节根据href地址获取相应
                for chapter_dict in book_info['catalogs']:
                    a_url = chapter_dict['a_href']
                    # 字典的get,方法有可能获得不到数据但是也不会报错
                    chapter_html = new_all_pages_dict.get(a_url)
                    # print(a_url)
                    del chapter_dict['a_href']
                    # print(chapter_html)
                    if chapter_html is not None:
                        if '您要查看的页面不存在或已删除' not in chapter_html:
                            chapter_html_obj = etree.HTML(chapter_html)
                            li_obj_list = chapter_html_obj.xpath('.//div[@class="Menubox"]/ul/li')
                            if li_obj_list:
                                for li_obj in li_obj_list:
                                    son_info_dict = {}
                                    name = li_obj.xpath('./text()')[0]
                                    son_info_dict['name'] = name
                                    son_info_dict['catalogs'] = []
                                    son_info_dict['pages'] = []
                                    li_id = li_obj.xpath('./@id')[0]
                                    img_short_url_list = chapter_html_obj.xpath('.//div[@id="con_one_{}"]//img/@src'.format(li_id[-1]))
                                    # print(img_short_url_list)
                                    img_pre_url = '/'.join(info_dict['url_4'].split('/')[:-1])
                                    if len(img_short_url_list) > 1:
                                        for img_short_url in img_short_url_list:
                                            img_url = os.path.join(img_pre_url, img_short_url).split("'")[1]
                                            # print(os.path.join(img_pre_url, img_short_url))
                                            if img_url not in chapter_dict['pages']:
                                                chapter_dict['pages'].append(img_url)
                                            son_info_dict['pages'].append(img_url)
                                            if img_url not in book_info['pages']:
                                                book_info['pages'].append(img_url)
                                    else:
                                        img_url = os.path.join(img_pre_url, img_short_url_list[0]).split("'")[1]
                                        son_info_dict['pages'].append(img_url)
                                        if img_url not in chapter_dict['pages']:
                                            chapter_dict['pages'].append(img_url)
                                        if img_url not in book_info['pages']:
                                            book_info['pages'].append(img_url)
                                    son_info_dict['pages'] = book_info['pages'].index(son_info_dict['pages'][0])

                                    chapter_dict['catalogs'].append(son_info_dict)

                                chapter_dict['pages'] = book_info['pages'].index(chapter_dict['pages'][0])
                        else:
                            chapter_dict['pages'] = None
                        # print(chapter_dict)
                    else:
                        chapter_dict['pages'] = None

                print(json.dumps(book_info, ensure_ascii=False))
                print(i)
                i += 1
                with open(self.current_path_name + '/english_171/analys_already_book.txt', 'a', encoding='utf-8') as f:
                    f.write(json.dumps(book_info, ensure_ascii=False)+'\n')
            except Exception:
                with open(self.current_path_name + '/english_171/error_analys_book.txt', 'a', encoding='utf-8') as f:
                    f.write(json.dumps(info_dict, ensure_ascii=False)+'\n')

    def error_analys_info(self):
        with open(self.current_path_name + '/english_171/error_analys_book.txt', 'r', encoding='utf-8') as f:
            # f.write(json.dumps(url_info_dict, ensure_ascii=False) + '\n')
            info_dict_str_list = f.readlines()
            print(len(info_dict_str_list))
        i = 0
        for info_dict_str in info_dict_str_list:
            info_dict = eval(info_dict_str)
            # try:
            print(info_dict['url_2'])
            book_info = {}
            chapter_html = info_dict['chapter_html']
            html_obj = etree.HTML(chapter_html)
            base_info = html_obj.xpath('.//div[@class="daohan"]//a/text()')
            if re.search(r'一年级|二年级|三年级|四年级|五年级|六年级', base_info[3]):
                book_info['period'] = '小学'
                book_info['grade'] = re.search(r'一年级|二年级|三年级|四年级|五年级|六年级', base_info[3]).group()
            elif re.search(r'七年级|八年级|九年级', base_info[3]):
                book_info['period'] = '初中'
                book_info['grade'] = re.search('七年级|八年级|九年级', base_info[3]).group()
            elif re.search(r'高一|高二|高三', base_info[3]):
                book_info['period'] = '高中'
                book_info['grade'] = re.search(r'高一|高二|高三', base_info[3]).group()
            book_info['press_version'] = base_info[1]
            book_info['subject'] = base_info[2]
            book_info['catalogs'] = []
            book_info['pages'] = []
            # print(book_info)
            # 提取章节列表
            a_obj_list = html_obj.xpath('.//div[@class="mululist"]//a')[1:]
            for a_obj in a_obj_list:
                name_list = a_obj.xpath('./text()|./*/text()')
                if len(name_list) > 1:
                    name = ' '.join(name_list)
                else:
                    name = name_list[0]
                a_href = 'http://www.dzkbw.com'+a_obj.xpath('./@href')[0]
                book_info['catalogs'].append({'name': name, 'catalogs': [], 'pages': [], 'a_href': a_href})
            # print(book_info)

            all_pages_list = info_dict['pages']
            # 把原始存储的列表字典，改成pages的字典，便于获得每个章节的html
            new_all_pages_dict = {}
            for page_dict in all_pages_list:
                for key, value in page_dict.items():
                    new_all_pages_dict[key] = value

            # 遍历每个章节根据href地址获取相应
            for chapter_dict in book_info['catalogs']:
                a_url = chapter_dict['a_href']
                # 字典的get,方法有可能获得不到数据但是也不会报错
                chapter_html = new_all_pages_dict.get(a_url)
                # print(a_url)
                del chapter_dict['a_href']
                # print(chapter_html)
                if chapter_html is not None:
                    if '您要查看的页面不存在或已删除' not in chapter_html:
                        chapter_html_obj = etree.HTML(chapter_html)
                        li_obj_list = chapter_html_obj.xpath('.//div[@class="Menubox"]/ul/li')
                        if li_obj_list:
                            for li_obj in li_obj_list:
                                son_info_dict = {}
                                name = li_obj.xpath('./text()')[0]
                                son_info_dict['name'] = name
                                son_info_dict['catalogs'] = []
                                son_info_dict['pages'] = []
                                li_id = li_obj.xpath('./@id')[0]
                                # 当img_short_url_list 为空的时候有孙子页码，提取链接
                                img_short_url_list = chapter_html_obj.xpath('.//div[@id="con_one_{}"]//img/@src'.format(li_id[-1]))
                                # print(img_short_url_list)
                                if img_short_url_list:
                                    img_pre_url = '/'.join(info_dict['url_4'].split('/')[:-1])
                                    if len(img_short_url_list) > 1:
                                        for img_short_url in img_short_url_list:
                                            img_url = os.path.join(img_pre_url, img_short_url).split("'")[1]
                                            # print(os.path.join(img_pre_url, img_short_url))
                                            if img_url not in chapter_dict['pages']:
                                                chapter_dict['pages'].append(img_url)
                                            son_info_dict['pages'].append(img_url)
                                            if img_url not in book_info['pages']:
                                                book_info['pages'].append(img_url)
                                    else:
                                        img_url = os.path.join(img_pre_url, img_short_url_list[0]).split("'")[1]
                                        son_info_dict['pages'].append(img_url)
                                        if img_url not in chapter_dict['pages']:
                                            chapter_dict['pages'].append(img_url)
                                        if img_url not in book_info['pages']:
                                            book_info['pages'].append(img_url)
                                    son_info_dict['pages'] = book_info['pages'].index(son_info_dict['pages'][0])

                                else:
                                    pre_url = '/'.join(info_dict['url_4'].split('/')[:-1])
                                    a_short_url = chapter_html_obj.xpath('.//div[@id="con_one_{}"]//a/@href'.format(li_id[-1]))[0]
                                    a_url = os.path.join(pre_url, a_short_url)
                                    res = self.send_request(a_url)
                                    html_obj = self.bytes_to_html(res)
                                    li_obj_list = html_obj.xpath('.//div[@class="Menubox"]/ul/li')
                                    for li_obj in li_obj_list:
                                        son_son_info_dict = {}
                                        name = li_obj.xpath('./text()')[0]
                                        son_son_info_dict['name'] = name
                                        son_son_info_dict['catalogs'] = []
                                        son_son_info_dict['pages'] = []
                                        li_id = li_obj.xpath('./@id')[0]
                                        img_short_url_list = chapter_html_obj.xpath('.//div[@id="con_one_{}"]//img/@src'.format(li_id[-1]))
                                        if len(img_short_url_list) > 1:
                                            for img_short in img_short_url_list:
                                                img_url = os.path.join(pre_url, img_short)
                                                if img_url not in son_son_info_dict['pages']:
                                                    son_son_info_dict['pages'].append(img_url)
                                                if img_url not in book_info['pages']:
                                                    book_info['pages'].append(img_url)

                                        son_son_info_dict['pages'] = book_info['pages'].index(son_son_info_dict['pages'][0])
                                        son_info_dict['catalogs'].append(son_son_info_dict)

                                chapter_dict['catalogs'].append(son_info_dict)

                            chapter_dict['pages'] = book_info['pages'].index(chapter_dict['pages'][0])
                    else:
                        chapter_dict['pages'] = None
                else:
                    chapter_dict['pages'] = None
                    # print(chapter_dict)

            print(json.dumps(book_info, ensure_ascii=False))
            print(i)
            i += 1
            # with open(self.current_path_name + '/english_171/analys_already_book.txt', 'a', encoding='utf-8') as f:
            #     f.write(json.dumps(book_info, ensure_ascii=False)+'\n')





if __name__ == '__main__':
    english = BookList171English()
    # english.get_book_list2()
    english.analys_info()
    # english.error_analys_info()
