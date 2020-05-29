import requests
import re
import os
import json
from retrying import retry
from lxml import etree
from requests import request
from openpyxl import load_workbook


class BookListYixue(object):

    def __init__(self):
        self.url_list = []
        self.headers = {
            'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3',
            'accept-encoding': 'gzip, deflate, br',
            'accept-language': 'zh-CN,zh;q=0.9',
            'cache-control': 'max-age=0',
            'cookie': 'PHPSESSID=h7gmm3nacj4t39ekoluj1rnm6v; Hm_lvt_8449acf5293ea0ed8a6ed77e29a4ba4d=1569729025; tgw_l7_route=abe66d4738da9ea03eeb99abf1385dd7; Hm_lpvt_8449acf5293ea0ed8a6ed77e29a4ba4d=1569741242',
            'upgrade-insecure-requests': '1',
            'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
        }
        self.headers_dzkb = {
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3',
            'Accept-Encoding': 'gzip, deflate',
            'Accept-Language': 'zh-CN,zh;q=0.9',
            'Connection': 'keep-alive',
            'Host': 'www.dzkbw.com',
            'Upgrade-Insecure-Requests': '1',
            'Cookie': 'ASPSESSIONIDAQSDDTCR=PNODCHKBCOCBGOFPAHEFELJN; _ga=GA1.2.1041419005.1569411852; _gid=GA1.2.1303393327.1569411852; Hm_lvt_463b13a51a26c92e0b49cd9e0dedb52b=1569411854; ASPSESSIONIDASSDCTDQ=LGDBIFLBPJEDOCCFNDDKEPEG; ASPSESSIONIDCSSADBCD=MPAAKLMBIGBFEBCLPNOJJFEM; ASPSESSIONIDCQQCDTCR=BCFFPGACFFKPBEOAABPIIAEN; ASPSESSIONIDCSRBARBS=FCIJEJACMOENGAABODEJJAEC; ASPSESSIONIDCQSBCACD=OAAAOIACBHBMFKPMDNJMOFGP; ASPSESSIONIDCSSADACD=HBGGNJBCGLJFNCKPFJCAHLHP; ASPSESSIONIDCSSACADC=KELCFACCHDIFGLOODAIMBPLP; ASPSESSIONIDASRDCSCR=CGEJILCCAJKCMBCMDMMMKFDD; ASPSESSIONIDASSBDADD=CJBGEDDCOJFGDENNNIBNNIBI; ASPSESSIONIDCSRCDSCQ=IBNOHEDCBCFCCAOEHPIKMGJN; ASPSESSIONIDCQRCBACD=NDOFENDCFJFCJNGMFANONOLD; ASPSESSIONIDCSTAAQBS=NFKAHODCHNABKLECPJKNLMGF; ASPSESSIONIDAQQBAQAT=IDNALKECICKPIABCCHCJOAFA; ASPSESSIONIDASRCDTDR=KPNCEMFCDOJMDKJPNAPKJKHB; ASPSESSIONIDAQSBCBDD=PAGBBBGCIMEHGBBCECHLHFKJ; ASPSESSIONIDASSDBTCR=FLJMCIGCHLJFEDAHJIIMLCOA; ASPSESSIONIDASRBARAT=IKKAJAHCFPEBHGKFKCECFLMH; ASPSESSIONIDCQSCCBCD=HJFFMDHCKPLNJHCCIIDLOLLD; ASPSESSIONIDCSSDCACD=MIDDBCLCPBENHPBOOOAJKMPK; ASPSESSIONIDASSBBQAT=JEDPFJLCIGNPAPJAIKFFNMHC; ASPSESSIONIDAQTDCBCC=LHIIDLLCEGNOJBPPKPPAMECH; ASPSESSIONIDASRCCTCQ=PANMOLLCPLJJGMFCEDANBIDD; ASPSESSIONIDAQTBADDD=ACKKLKMCDEEDFDPKCJIALFMA; ASPSESSIONIDASRDCBDC=NJIKLFNCDLEKDMCNNBHJPFON; ASPSESSIONIDCSTCBBCD=KLMDOMNCNPNAMGKGPPJECOFL; _gat_gtag_UA_66838525_1=1; Hm_lpvt_463b13a51a26c92e0b49cd9e0dedb52b=1569569659',
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.75 Safari/537.36'
        }
        self.current_path_name = os.path.dirname(os.path.abspath(__file__))
    @retry(stop_max_attempt_number=6, stop_max_delay=3000, wait_fixed=3000, wait_incrementing_increment=500)
    def send_request(self, url, headers=None):
        requests.packages.urllib3.disable_warnings()
        res = request(method='get', url=url, headers=headers, timeout=20, verify=False)
        return res.content

    def create_file_path(self):
        if not os.path.exists(self.current_path_name+'/yixue'):
            os.makedirs(self.current_path_name+'/yixue')

    def struct_url(self):
        open_excel = load_workbook(self.current_path_name+'/book_list.xlsx')
        tab = open_excel['book_list']
        max_num = tab.max_row
        for i in range(1, max_num+1):
            url_1 = tab.cell(i, 1).value
            url_2 = tab.cell(i, 2).value
            url_3 = tab.cell(i, 3).value
            url_4 = tab.cell(i, 4).value
            book_url_obj = re.search('https://www.yixuela.com/books.*isstop=True', url_4)
            if book_url_obj:
                self.url_list.append({'url_1': url_1, 'url_2': url_2, 'url_3': url_3, 'url_4': url_4})

    # 把字节转成字符串
    def bytes_to_html(self, html_bytes):
        try:
            html_obj = etree.HTML(html_bytes.decode(encoding='utf-8'))
        except Exception:
            try:
                html_obj = etree.HTML(html_bytes.decode(encoding='gbk'))
            except Exception:
                try:
                    html_obj = etree.HTML(html_bytes.decode(encoding='GB2312'))
                except Exception:
                    html_obj = etree.HTML(html_bytes.decode(encoding='GB18030'))
        return html_obj

    # 转换成html字符串
    def bytes_to_str(self, html_bytes):
        try:
            html_obj = html_bytes.decode(encoding='utf-8')
        except Exception:
            try:
                html_obj = html_bytes.decode(encoding='gbk')
            except Exception:
                try:
                    html_obj = html_bytes.decode(encoding='GB2312')
                except Exception:
                    html_obj = html_bytes.decode(encoding='GB18030')
        return html_obj

    def anay_subj(self, html_bytes):
        html_obj = self.bytes_to_html(html_bytes)
        info_list = html_obj.xpath('.//div[@class="daohan"]/*/text()')
        return info_list

    # 解析书本的目录列表
    def analys_book_html(self, html_bytes):
        html_obj = self.bytes_to_html(html_bytes)
        a_obj_list = html_obj.xpath('.//nav[@class="right-nav"]/ul/li/a')
        url_list = []
        for a_obj in a_obj_list:
            url = 'https://www.yixuela.com'+a_obj.xpath('./@href')[0]
            url_list.append(url)
        return url_list

    def analys_img(self, html_bytes):
        # import pdb
        # pdb.set_trace()
        html_obj = self.bytes_to_str(html_bytes)
        img_url_list = html_obj.xpath('.//div[@class="book-imgs-list bg-white pad-15"]/img/@src')
        return img_url_list


    def run(self):
        self.create_file_path()
        self.struct_url()
        print(len(self.url_list))
        for url_info_dict in self.url_list:
            url_str = url_info_dict['url_4']
            book_url = re.search('https://www.yixuela.com/books.*isstop=True', url_str).group()
            print('*'*20)
            print(book_url)
            try:
                # 获得原始目录的html
                star_url = url_info_dict['url_1'].replace('\n', '')
                html_bytes = self.send_request(url=star_url, headers=self.headers_dzkb)
                url_info_dict['chapter_html'] = self.bytes_to_str(html_bytes)
                # 获得书本的html
                book_html_bytes = self.send_request(url=book_url, headers=self.headers)
                # 解析出书本的目录地址
                mulu_url_list = self.analys_book_html(book_html_bytes)
                page_dict = {}
                for url_sub in mulu_url_list:
                    print(url_sub)
                    res_obj = self.send_request(url=url_sub, headers=self.headers)
                    html_str = self.bytes_to_str(res_obj)
                    page_dict[url_sub] = html_str
                url_info_dict['pages'] = page_dict
                with open(self.current_path_name+'/yixue/already.txt', 'a+', encoding='utf-8') as f:
                    f.write(json.dumps(url_info_dict)+'\n')
            except Exception:
                with open(self.current_path_name+'/yixue/error_url.txt', 'a+', encoding='utf-8') as f_error:
                    f_error.write(json.dumps(url_info_dict)+'\n')


if __name__ == '__main__':
    booklistyixue = BookListYixue()
    booklistyixue.run()
