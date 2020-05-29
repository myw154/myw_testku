import requests
import re
import os
import json
from retrying import retry
from lxml import etree
from requests import request
from openpyxl import load_workbook


class BookListShuJiu(object):

    def __init__(self):
        self.url_list = []
        self.headers = {
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3',
            'Accept-Encoding': 'gzip, deflate',
            'Accept-Language': 'zh-CN,zh;q=0.9',
            'Cache-Control': 'max-age=0',
            'Connection': 'keep-alive',
            'Cookie': 'ASPSESSIONIDCCSSTBRR=IHJOOLNDJBPEMCIJMLJGDNLC; __51cke__=; ASPSESSIONIDCATRRCRR=GDDFBIKAEPONFMGAEENMFOGJ; __tins__19358686=%7B%22sid%22%3A%201569810431058%2C%20%22vd%22%3A%2027%2C%20%22expires%22%3A%201569813141141%7D; __51laig__=208',
            'Host': 'www.shuxue9.com',
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36'
        }
        self.headers_dzkb = {
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3',
            'Accept-Encoding': 'gzip, deflate',
            'Accept-Language': 'zh-CN,zh;q=0.9',
            'Connection': 'keep-alive',
            'Host': 'www.dzkbw.com',
            'Upgrade-Insecure-Requests': '1',
            'Cookie': 'ASPSESSIONIDAQSDDTCR=PNODCHKBCOCBGOFPAHEFELJN; _ga=GA1.2.1041419005.1569411852; _gid=GA1.2.1303393327.1569411852; Hm_lvt_463b13a51a26c92e0b49cd9e0dedb52b=1569411854; ASPSESSIONIDASSDCTDQ=LGDBIFLBPJEDOCCFNDDKEPEG; ASPSESSIONIDCSSADBCD=MPAAKLMBIGBFEBCLPNOJJFEM; ASPSESSIONIDCQQCDTCR=BCFFPGACFFKPBEOAABPIIAEN; ASPSESSIONIDCSRBARBS=FCIJEJACMOENGAABODEJJAEC; ASPSESSIONIDCQSBCACD=OAAAOIACBHBMFKPMDNJMOFGP; ASPSESSIONIDCSSADACD=HBGGNJBCGLJFNCKPFJCAHLHP; ASPSESSIONIDCSSACADC=KELCFACCHDIFGLOODAIMBPLP; ASPSESSIONIDASRDCSCR=CGEJILCCAJKCMBCMDMMMKFDD; ASPSESSIONIDASSBDADD=CJBGEDDCOJFGDENNNIBNNIBI; ASPSESSIONIDCSRCDSCQ=IBNOHEDCBCFCCAOEHPIKMGJN; ASPSESSIONIDCQRCBACD=NDOFENDCFJFCJNGMFANONOLD; ASPSESSIONIDCSTAAQBS=NFKAHODCHNABKLECPJKNLMGF; ASPSESSIONIDAQQBAQAT=IDNALKECICKPIABCCHCJOAFA; ASPSESSIONIDASRCDTDR=KPNCEMFCDOJMDKJPNAPKJKHB; ASPSESSIONIDAQSBCBDD=PAGBBBGCIMEHGBBCECHLHFKJ; ASPSESSIONIDASSDBTCR=FLJMCIGCHLJFEDAHJIIMLCOA; ASPSESSIONIDASRBARAT=IKKAJAHCFPEBHGKFKCECFLMH; ASPSESSIONIDCQSCCBCD=HJFFMDHCKPLNJHCCIIDLOLLD; ASPSESSIONIDCSSDCACD=MIDDBCLCPBENHPBOOOAJKMPK; ASPSESSIONIDASSBBQAT=JEDPFJLCIGNPAPJAIKFFNMHC; ASPSESSIONIDAQTDCBCC=LHIIDLLCEGNOJBPPKPPAMECH; ASPSESSIONIDASRCCTCQ=PANMOLLCPLJJGMFCEDANBIDD; ASPSESSIONIDAQTBADDD=ACKKLKMCDEEDFDPKCJIALFMA; ASPSESSIONIDASRDCBDC=NJIKLFNCDLEKDMCNNBHJPFON; ASPSESSIONIDCSTCBBCD=KLMDOMNCNPNAMGKGPPJECOFL; _gat_gtag_UA_66838525_1=1; Hm_lpvt_463b13a51a26c92e0b49cd9e0dedb52b=1569569659',
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.75 Safari/537.36'
        }
        # /Users/guo/Desktop/mywtest/xuezi_net
        self.current_path_name = os.path.dirname(os.path.abspath(__file__))

    def create_file_path(self):
        if not os.path.exists(self.current_path_name+'/shujiu'):
            os.makedirs(self.current_path_name+'/shujiu')
            os.makedirs(self.current_path_name+'/shujiu/img')
            os.makedirs(self.current_path_name+'/shujiu/chapter_html')

    @retry(stop_max_attempt_number=6, stop_max_delay=3000, wait_fixed=3000, wait_incrementing_increment=500)
    def send_request(self, url, headers=None):
        requests.packages.urllib3.disable_warnings()
        res = request(method='get', url=url, headers=headers, timeout=20, verify=False)
        return res.content

    def struct_url(self):
        open_excel = load_workbook(os.path.abspath('book_list.xlsx'))
        tab =open_excel['book_list']
        max_num = tab.max_row
        #for i in range(1, max_num+1):
        for i in range(437, 437+1):
            url_1 = tab.cell(i, 1).value
            url_2 = tab.cell(i, 2).value
            url_3 = tab.cell(i, 3).value
            url_4 = tab.cell(i, 4).value
            url_re_obj = re.search(r'http://www.shuxue9.com/(.+)isstop=True', url_4)
            if url_re_obj:
                self.url_list.append({'url_1': url_1, 'url_2': url_2, 'url_3': url_3, 'url_4': url_4})

    # 把字节转成字符串
    def bytes_to_obj(self, html_bytes):
        try:
            html_obj = etree.HTML(html_bytes.decode(encoding='utf-8'))
        except Exception:
            try:
                html_obj = etree.HTML(html_bytes.decode(encoding='gbk'))
            except Exception:
                try:
                    html_obj = etree.HTML(html_bytes.decode(encoding='GB2312'))
                except Exception:
                    html_obj = etree.HTML(html_bytes.decode(encoding='GB18030'))
        return html_obj

    # 转换成html字符串
    def bytes_to_str(self, html_bytes):
        try:
            html_obj = html_bytes.decode(encoding='utf-8')
        except Exception:
            try:
                html_obj = html_bytes.decode(encoding='gbk')
            except Exception:
                try:
                    html_obj = html_bytes.decode(encoding='GB2312')
                except Exception:
                    html_obj = html_bytes.decode(encoding='GB18030')
        return html_obj

    def anay_subj(self, html_bytes):
        html_obj = self.bytes_to_str(html_bytes)
        info_list = html_obj.xpath('.//div[@class="daohan"]/*/text()')
        return info_list

    # 分析有多少页 返回总页数
    def analys_chapter_num(self, html_bytes):
        html_obj = self.bytes_to_str(html_bytes)
        # html_str = etree.tostring(html_obj).decode()
        page_re_obj = re.search(r'tpage=(\d+)', html_obj)
        page_num = int(page_re_obj.group(1))
        return page_num

    def get_book(self, url):
        print(url)
        # 获得书本的html
        import pdb
        pdb.set_trace()
        book_html_bytes = self.send_request(url, self.headers)
        page_num = self.analys_chapter_num(book_html_bytes)
        return page_num

    def get_book_num(self, boo_url):
        res_obj = self.send_request(boo_url)
        html_str = self.bytes_to_str(res_obj)
        page_num = int(re.search(r'maxPage=(\d+);', html_str).group(1))
        return page_num

    def get_all_chapter(self, url, page_num):
        pages_dict = {}
        for i in range(1, page_num+1):
            chapter_url = re.sub(r'/\d+.html', '/{}.html'.format(i), url)
            print(chapter_url)
            res_obj = self.send_request(chapter_url)
            html_str = self.bytes_to_str(res_obj)
            pages_dict[chapter_url] = html_str
        return pages_dict

    def run(self):
        #构造请求的地址
        self.create_file_path()
        self.struct_url()
        print(len(self.url_list))
        for url_info_dict in self.url_list:
            book_url = url_info_dict['url_4'].split("'")[1]

            # 获得原始目录的html
            try:
                star_url = url_info_dict['url_1']
                html_bytes = self.send_request(url=star_url.replace('\n', ''), headers=self.headers_dzkb)
                chapter_html = self.bytes_to_str(html_bytes)
                # 供提取学科版本用
                url_info_dict['chapter_html'] = chapter_html
                # 获得这本书
                page_num = self.get_book_num(url_info_dict['url_2'])
                url_info_dict['page_all_num'] = page_num
                print('*'*20)
                print(book_url)
                print(page_num)
                chapter_pages = self.get_all_chapter(book_url, page_num)
                url_info_dict['pages'] = chapter_pages
                with open(self.current_path_name + '/shujiu/already_url.txt', 'a+', encoding='utf-8') as f_file:
                    f_file.write(json.dumps(url_info_dict, ensure_ascii=False)+'\n')
            except Exception:
                with open(self.current_path_name + '/shujiu/error_url.txt', 'a+', encoding='utf8') as f_error:
                     f_error.write(json.dumps(url_info_dict)+'\n')


if __name__ == '__main__':
    booklistshujiu = BookListShuJiu()
    booklistshujiu.run()
