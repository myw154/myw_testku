import requests
import re
import os
import json
from retrying import retry
from lxml import etree
from requests import request
from openpyxl import load_workbook


class BookListPep(object):

    def __init__(self):
        self.url_list = []
        self.book_url = []
        self.book_first_url = []
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36'
        }
        self.headers_dzkb = {
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3',
            'Accept-Encoding': 'gzip, deflate',
            'Accept-Language': 'zh-CN,zh;q=0.9',
            'Connection': 'keep-alive',
            'Host': 'www.dzkbw.com',
            'Upgrade-Insecure-Requests': '1',
            'Cookie': 'ASPSESSIONIDAQSDDTCR=PNODCHKBCOCBGOFPAHEFELJN; _ga=GA1.2.1041419005.1569411852; _gid=GA1.2.1303393327.1569411852; Hm_lvt_463b13a51a26c92e0b49cd9e0dedb52b=1569411854; ASPSESSIONIDASSDCTDQ=LGDBIFLBPJEDOCCFNDDKEPEG; ASPSESSIONIDCSSADBCD=MPAAKLMBIGBFEBCLPNOJJFEM; ASPSESSIONIDCQQCDTCR=BCFFPGACFFKPBEOAABPIIAEN; ASPSESSIONIDCSRBARBS=FCIJEJACMOENGAABODEJJAEC; ASPSESSIONIDCQSBCACD=OAAAOIACBHBMFKPMDNJMOFGP; ASPSESSIONIDCSSADACD=HBGGNJBCGLJFNCKPFJCAHLHP; ASPSESSIONIDCSSACADC=KELCFACCHDIFGLOODAIMBPLP; ASPSESSIONIDASRDCSCR=CGEJILCCAJKCMBCMDMMMKFDD; ASPSESSIONIDASSBDADD=CJBGEDDCOJFGDENNNIBNNIBI; ASPSESSIONIDCSRCDSCQ=IBNOHEDCBCFCCAOEHPIKMGJN; ASPSESSIONIDCQRCBACD=NDOFENDCFJFCJNGMFANONOLD; ASPSESSIONIDCSTAAQBS=NFKAHODCHNABKLECPJKNLMGF; ASPSESSIONIDAQQBAQAT=IDNALKECICKPIABCCHCJOAFA; ASPSESSIONIDASRCDTDR=KPNCEMFCDOJMDKJPNAPKJKHB; ASPSESSIONIDAQSBCBDD=PAGBBBGCIMEHGBBCECHLHFKJ; ASPSESSIONIDASSDBTCR=FLJMCIGCHLJFEDAHJIIMLCOA; ASPSESSIONIDASRBARAT=IKKAJAHCFPEBHGKFKCECFLMH; ASPSESSIONIDCQSCCBCD=HJFFMDHCKPLNJHCCIIDLOLLD; ASPSESSIONIDCSSDCACD=MIDDBCLCPBENHPBOOOAJKMPK; ASPSESSIONIDASSBBQAT=JEDPFJLCIGNPAPJAIKFFNMHC; ASPSESSIONIDAQTDCBCC=LHIIDLLCEGNOJBPPKPPAMECH; ASPSESSIONIDASRCCTCQ=PANMOLLCPLJJGMFCEDANBIDD; ASPSESSIONIDAQTBADDD=ACKKLKMCDEEDFDPKCJIALFMA; ASPSESSIONIDASRDCBDC=NJIKLFNCDLEKDMCNNBHJPFON; ASPSESSIONIDCSTCBBCD=KLMDOMNCNPNAMGKGPPJECOFL; _gat_gtag_UA_66838525_1=1; Hm_lpvt_463b13a51a26c92e0b49cd9e0dedb52b=1569569659',
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.75 Safari/537.36'
        }
        # /Users/guo/Desktop/mywtest/xuezi_net
        self.current_path_name = os.path.dirname(os.path.abspath(__file__))

    def create_file_path(self):
        if not os.path.exists(self.current_path_name+'/pep'):
            os.makedirs(self.current_path_name+'/pep')
        if not os.path.exists(self.current_path_name+'/pep/img'):
            os.makedirs(self.current_path_name+'/pep/img')
        if not os.path.exists(self.current_path_name+'/pep/book_chapter_html'):
            os.makedirs(self.current_path_name+'/pep/book_chapter_html')
        if not os.path.exists(self.current_path_name + '/pep/original_chapter_html/'):
            os.makedirs(self.current_path_name + '/pep/original_chapter_html/')

    @retry(stop_max_attempt_number=5, stop_max_delay=3000, wait_fixed=3000, wait_incrementing_increment=500)
    def send_request(self, url, headers=None):
        requests.packages.urllib3.disable_warnings()
        res = request(method='get', url=url, headers=headers, timeout=10, verify=False)
        return res.content

    def struct_url(self):
        open_excel = load_workbook(os.path.abspath('book_list.xlsx'))
        tab = open_excel['book_list']
        max_num = tab.max_row
        # for i in range(225, 227):
        for i in range(1, max_num+1):
            url_1 = tab.cell(i, 1).value
            url_2 = tab.cell(i, 2).value
            url_3 = tab.cell(i, 3).value
            url_4 = tab.cell(i, 4).value
            re_obj = re.search(r'http://www.pep.com.cn/(.+)isstop=True', url_4)
            if re_obj:
                self.book_url.append({'url_1': url_1, 'url_2': url_2, 'url_3': url_3, 'url_4': url_4})

    # 把字节转成字符串
    def bytes_to_html(self, html_bytes):
        try:
            html_obj = etree.HTML(html_bytes.decode(encoding='utf-8'))
        except Exception:
            try:
                html_obj = etree.HTML(html_bytes.decode(encoding='gbk'))
            except Exception:
                try:
                    html_obj = etree.HTML(html_bytes.decode(encoding='GB2312'))
                except Exception:
                    html_obj = etree.HTML(html_bytes.decode(encoding='GB18030'))
        return html_obj

    # 转换成html字符串
    def bytes_to_str(self, html_bytes):
        try:
            html_obj = html_bytes.decode(encoding='utf-8')
        except Exception:
            try:
                html_obj = html_bytes.decode(encoding='gbk')
            except Exception:
                try:
                    html_obj = html_bytes.decode(encoding='GB2312')
                except Exception:
                    html_obj = html_bytes.decode(encoding='GB18030')
        return html_obj

    def anay_subj(self, html_bytes):
        html_obj = self.bytes_to_html(html_bytes)
        info_list = html_obj.xpath('.//div[@class="daohan"]/*/text()')
        # 获得原始书本的章节目录
        mulu_list = []
        mulu_a_list = html_obj.xpath('.//div[@class="bookmulu"]/a')
        for a_obj in mulu_a_list:
            name = a_obj.xpath('./*/text() | ./text()')[0] if len(a_obj.xpath('./*/text() | ./text()')) > 0 else None
            href = 'http://www.dzkbw.com'+a_obj.xpath('./@href')[0]
            mulu_list.append([name, href])
        return info_list, mulu_list

    # 构造每一本书的请求地址
    def struct_book_url(self):
        # 创建一些必要的文件夹路径
        self.create_file_path()
        # 构造书本的地址
        self.struct_url()

    # 下载图片
    def download_img(self, book_url, chapter_url, img_url):
        book_name = '_'.join(book_url.split('/')[3:-1])
        if not os.path.exists(self.current_path_name+'/pep/img/'+book_name):
            os.makedirs(self.current_path_name+'/pep/img/'+book_name)
        if chapter_url == img_url:
            img_name = '_'.join(img_url.split('/')[3:]).replace('?isstop=True', '')
        else:
            img_name = '_'.join(img_url.split('/')[3:])
        try:
            img_path = self.current_path_name+'/pep/img/'+book_name+'/'+img_name
            res_bytes = self.send_request(img_url)
            with open(img_path, 'wb') as f:
                f.write(res_bytes)
            with open(self.current_path_name + '/pep/book_chapter_img_url.txt', 'a', encoding='utf-8') as f:
                f.write(str([book_url, chapter_url, img_url])+'\n')
        except Exception:
            with open(self.current_path_name+'/pep/img_url_error.txt', 'a', encoding='utf-8') as f:
                f.write(str([book_url, chapter_url, img_url])+'\n')

    # 获得书本的目录信息
    def get_new_mulu(self, book_url):
        res_bytes = self.send_request(book_url, self.headers_dzkb)
        html_obj = self.bytes_to_str(res_bytes)
        page_num = int(re.search('maxPage=(\d+);', html_obj).group(1))
        return html_obj, page_num

    # 获得一个章节的新的地址
    def get_new_url(self, url):
        res_bytes = self.send_request(url, self.headers_dzkb)
        html_obj = self.bytes_to_html(res_bytes)
        # 获得二维码的地址
        if html_obj.xpath('.//div[@class="bookimg"]//a[2]/@href'):
            if not html_obj.xpath('.//div[@class="bookimg"]//a[2]/@href')[0].startswith('/go.asp?'):
                seract_url = 'http://www.dzkbw.com'+html_obj.xpath('.//div[@class="bookimg"]//a[1]/@href')[0]
            else:
                seract_url = 'http://www.dzkbw.com' + html_obj.xpath('.//div[@class="bookimg"]//a[2]/@href')[0]
        else:
            seract_url = 'http://www.dzkbw.com' + html_obj.xpath('.//div[@class="bookimg"]//a[1]/@href')[0]
        seract_res_bytes = self.send_request(seract_url, self.headers_dzkb)
        seract_obj_html = self.bytes_to_html(seract_res_bytes)
        new_chapter_url = seract_obj_html.xpath('.//head/meta/@content')[1].split("'")[1]
        return new_chapter_url

    # 获得每个章节的信息地址，并获得新信息
    def get_chapter(self, book_url, new_mulu_list):
        for a_href_list in new_mulu_list:
            try:
                # 获得每个章节的新地址
                new_url = self.get_new_url(a_href_list[0])
                # 如果新地址中含有.jpg 或者oldimgages 说明返回是图片地址
                if 'oldimages' in new_url or '.jpg' in new_url:
                    self.download_img(book_url, new_url, new_url)
                else:
                    self.get_content(book_url, new_url)
            except Exception:
                with open(self.current_path_name+'/pep/chapter_url_error.txt', 'a', encoding='utf-8') as f:
                    f.write(str([book_url, a_href_list[0]]))

    def get_content_old(self, book_url, new_chaper_url):
        res_bytes = self.send_request(new_chaper_url, self.headers)
        html_obj = self.bytes_to_html(res_bytes)
        img_short_url_list = html_obj.xpath('.//div[@id="doccontent"]//img/@src')
        for img_short_url in img_short_url_list:
            img_url = '/'.join(new_chaper_url.split('/')[:-1])+img_short_url.replace('./', '/')
            self.download_img(book_url, new_chaper_url, img_url)

    def get_content(self, new_chaper_url):
        res_bytes = self.send_request(new_chaper_url, self.headers)
        html_obj = self.bytes_to_str(res_bytes)
        return html_obj

    def get_all_chapter(self, book_url, page_num):
        pages_dict = {}
        for i in range(1, page_num+1):
            if i <= 9:
                chapter_url = re.sub(r'/\d+.htm', '/00'+str(i)+'.htm', book_url)
            elif 10 <= i <= 99:
                chapter_url = re.sub(r'/\d+.htm', '/0'+str(i)+'.htm', book_url)
            else:
                chapter_url = re.sub(r'/\d+.htm', '/'+str(i)+'.htm', book_url)

            new_url = self.get_new_url(chapter_url)
            print(chapter_url)
            # 如果新地址中含有.jpg 或者oldimgages 说明返回是图片地址
            if 'oldimages' in new_url or '.jpg' in new_url:
                pages_dict[chapter_url] = new_url
            else:
                html_str = self.get_content(new_url)
                pages_dict[chapter_url] = html_str
        return pages_dict

    # 获得每本书的目录信息，获得每本书的第一页的url地址
    def get_book_list(self):
        self.struct_book_url()
        print(len(self.book_url))
        for url_info_dict in self.book_url:
            # 获得原始目录的html
            try:
                start_url = url_info_dict['url_1']
                res_obj = self.send_request(start_url)
                chapter_html = self.bytes_to_str(res_obj)
                url_info_dict['chapter_html'] = chapter_html
                book_url = url_info_dict['url_2']
                # 获得书本的新的目录信息
                print(book_url)
                html_str, page_num = self.get_new_mulu(book_url)
                print(page_num)
                url_info_dict['chapter_html_2'] = html_str
                url_info_dict['pages_all_num'] = page_num
                pages_dict = self.get_all_chapter(book_url, page_num)
                url_info_dict['pages'] = pages_dict
                # 保存已经获得的信息
                with open(self.current_path_name+'/pep/already_get_book.txt', 'a+', encoding='utf-8') as f:
                    f.write(json.dumps(url_info_dict, ensure_ascii=False)+'\n')
            except Exception:
                with open(self.current_path_name+'/pep/error_url.txt', 'a+', encoding='utf-8') as f:
                    f.write(json.dumps(url_info_dict, ensure_ascii=False)+'\n')

    def get_error_url(self):
        with open(self.current_path_name + '/pep/error_url.txt', 'r', encoding='utf-8') as f:
            # f.write(json.dumps(url_info_dict, ensure_ascii=False) + '\n')
            book_url_dict_list = f.readlines()
            print(len(book_url_dict_list))
            for book_url_dict_str in book_url_dict_list:
                url_info_dict = eval(book_url_dict_str)
                try:
                    book_url = url_info_dict['url_2']
                    # 获得书本的新的目录信息
                    print('*'*20)
                    print(book_url)
                    html_str, page_num = self.get_new_mulu(book_url)
                    print(page_num)
                    url_info_dict['chapter_html'] = html_str
                    # 根据章节的地址构建 每个章节的起始和终止
                    url_info_dict['pages_all_num'] = page_num
                    pages_dict = self.get_all_chapter(book_url, page_num)
                    url_info_dict['pages'] = pages_dict
                    # 保存已经获得的信息
                    with open(self.current_path_name + '/pep/already_get_book.txt', 'a+', encoding='utf-8') as f:
                        f.write(json.dumps(url_info_dict, ensure_ascii=False) + '\n')
                except Exception:
                    with open(self.current_path_name + '/pep/error_url2.txt', 'a+', encoding='utf-8') as f:
                        f.write(json.dumps(url_info_dict, ensure_ascii=False) + '\n')

if __name__ == '__main__':
    pep = BookListPep()
    pep.get_book_list()
    #pep.get_error_url()
    # res_url = pep.get_new_url('http://www.dzkbw.com/books/rjb/yuwen/10s/100.htm')
    # print(res_url)








