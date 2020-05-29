import requests
import re
import os
import json
from retrying import retry
from lxml import etree
from requests import request
from openpyxl import load_workbook


class BookListRuiWen(object):

    def __init__(self):
        self.url_list = []
        self.book_url = []
        self.book_first_url = []
        self.headers = {
            'cookie': 'Hm_lvt_4d9c0095d55f98245e83c4b5879def56=1571370647,1571370718,1571371962,1571372000; Hm_lvt_ae653be27558292a41da382819e14a2f=1571380127; Hm_lpvt_ae653be27558292a41da382819e14a2f=1571380146; Hm_lpvt_4d9c0095d55f98245e83c4b5879def56=1571394887',
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36'
        }

        self.headers_dzkb = {
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3',
            'Accept-Encoding': 'gzip, deflate',
            'Accept-Language': 'zh-CN,zh;q=0.9',
            'Connection': 'keep-alive',
            'Host': 'www.dzkbw.com',
            'Upgrade-Insecure-Requests': '1',
            'Cookie': 'ASPSESSIONIDAQSDDTCR=PNODCHKBCOCBGOFPAHEFELJN; _ga=GA1.2.1041419005.1569411852; _gid=GA1.2.1303393327.1569411852; Hm_lvt_463b13a51a26c92e0b49cd9e0dedb52b=1569411854; ASPSESSIONIDASSDCTDQ=LGDBIFLBPJEDOCCFNDDKEPEG; ASPSESSIONIDCSSADBCD=MPAAKLMBIGBFEBCLPNOJJFEM; ASPSESSIONIDCQQCDTCR=BCFFPGACFFKPBEOAABPIIAEN; ASPSESSIONIDCSRBARBS=FCIJEJACMOENGAABODEJJAEC; ASPSESSIONIDCQSBCACD=OAAAOIACBHBMFKPMDNJMOFGP; ASPSESSIONIDCSSADACD=HBGGNJBCGLJFNCKPFJCAHLHP; ASPSESSIONIDCSSACADC=KELCFACCHDIFGLOODAIMBPLP; ASPSESSIONIDASRDCSCR=CGEJILCCAJKCMBCMDMMMKFDD; ASPSESSIONIDASSBDADD=CJBGEDDCOJFGDENNNIBNNIBI; ASPSESSIONIDCSRCDSCQ=IBNOHEDCBCFCCAOEHPIKMGJN; ASPSESSIONIDCQRCBACD=NDOFENDCFJFCJNGMFANONOLD; ASPSESSIONIDCSTAAQBS=NFKAHODCHNABKLECPJKNLMGF; ASPSESSIONIDAQQBAQAT=IDNALKECICKPIABCCHCJOAFA; ASPSESSIONIDASRCDTDR=KPNCEMFCDOJMDKJPNAPKJKHB; ASPSESSIONIDAQSBCBDD=PAGBBBGCIMEHGBBCECHLHFKJ; ASPSESSIONIDASSDBTCR=FLJMCIGCHLJFEDAHJIIMLCOA; ASPSESSIONIDASRBARAT=IKKAJAHCFPEBHGKFKCECFLMH; ASPSESSIONIDCQSCCBCD=HJFFMDHCKPLNJHCCIIDLOLLD; ASPSESSIONIDCSSDCACD=MIDDBCLCPBENHPBOOOAJKMPK; ASPSESSIONIDASSBBQAT=JEDPFJLCIGNPAPJAIKFFNMHC; ASPSESSIONIDAQTDCBCC=LHIIDLLCEGNOJBPPKPPAMECH; ASPSESSIONIDASRCCTCQ=PANMOLLCPLJJGMFCEDANBIDD; ASPSESSIONIDAQTBADDD=ACKKLKMCDEEDFDPKCJIALFMA; ASPSESSIONIDASRDCBDC=NJIKLFNCDLEKDMCNNBHJPFON; ASPSESSIONIDCSTCBBCD=KLMDOMNCNPNAMGKGPPJECOFL; _gat_gtag_UA_66838525_1=1; Hm_lpvt_463b13a51a26c92e0b49cd9e0dedb52b=1569569659',
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.75 Safari/537.36'
        }
        # /Users/guo/Desktop/mywtest/xuezi_net
        self.current_path_name = os.path.dirname(os.path.abspath(__file__))

    def create_file_path(self):
        if not os.path.exists(self.current_path_name+'/ruiwen'):
            os.makedirs(self.current_path_name+'/ruiwen')
        if not os.path.exists(self.current_path_name+'/ruiwen/img'):
            os.makedirs(self.current_path_name+'/ruiwen/img')
        if not os.path.exists(self.current_path_name+'/ruiwen/book_chapter_html'):
            os.makedirs(self.current_path_name+'/ruiwen/book_chapter_html')
        if not os.path.exists(self.current_path_name + '/ruiwen/original_chapter_html/'):
            os.makedirs(self.current_path_name + '/ruiwen/original_chapter_html/')

    @retry(stop_max_attempt_number=5, stop_max_delay=3000, wait_fixed=3000, wait_incrementing_increment=500)
    def send_request(self, url, headers=None):
        requests.packages.urllib3.disable_warnings()
        res = request(method='get', url=url, headers=headers, timeout=10, verify=False)
        return res

    def struct_url(self):
        open_excel = load_workbook(os.path.abspath('book_list.xlsx'))
        tab = open_excel['book_list']
        max_num = tab.max_row
        # for i in range(888, 889):
        for i in range(1, max_num+1):
            url_1 = tab.cell(i, 1).value
            url_2 = tab.cell(i, 2).value
            url_3 = tab.cell(i, 3).value
            url_4 = tab.cell(i, 4).value
            re_obj = re.search(r'https://www.ruiwen.com/(.*)', url_4)
            if re_obj:
                self.book_url.append({'url_1': url_1, 'url_2': url_2, 'url_3': url_3, 'url_4': url_4})

    # 把字节转成字符串
    def bytes_to_html(self, html_bytes):
        try:
            html_obj = etree.HTML(html_bytes.decode(encoding='utf-8'))
        except Exception:
            try:
                html_obj = etree.HTML(html_bytes.decode(encoding='gbk'))
            except Exception:
                try:
                    html_obj = etree.HTML(html_bytes.decode(encoding='GB2312'))
                except Exception:
                    html_obj = etree.HTML(html_bytes.decode(encoding='GB18030'))
        return html_obj

    # 转换成html字符串
    def bytes_to_str(self, html_bytes):
        try:
            html_obj = html_bytes.decode(encoding='utf-8')
        except Exception:
            try:
                html_obj = html_bytes.decode(encoding='gbk')
            except Exception:
                try:
                    html_obj = html_bytes.decode(encoding='GB2312')
                except Exception:
                    html_obj = html_bytes.decode(encoding='GB18030')
        return html_obj

    def anay_subj(self, html_bytes):
        html_obj = self.bytes_to_html(html_bytes.content)
        info_list = html_obj.xpath('.//div[@class="daohan"]/*/text()')
        # 获得原始书本的章节目录
        mulu_list = []
        mulu_a_list = html_obj.xpath('.//div[@class="bookmulu"]/a')
        for a_obj in mulu_a_list:
            name = a_obj.xpath('./*/text() | ./text()')[0] if len(a_obj.xpath('./*/text() | ./text()')) > 0 else None
            href = 'http://www.dzkbw.com'+a_obj.xpath('./@href')[0]
            mulu_list.append([name, href])
        return info_list, mulu_list

    # 构造每一本书的请求地址
    def struct_book_url(self):
        # 创建一些必要的文件夹路径
        self.create_file_path()
        # 构造书本的地址
        self.struct_url()

    def get_new_chapter_list(self, book_url):
        res_bytes = self.send_request(book_url.replace('?isstop=True', '').replace('https', 'http'), self.headers)
        html_obj = self.bytes_to_html(res_bytes.content)
        # 获得页码数
        page_num = re.search(r'\d+', html_obj.xpath('.//div[@class="right_box"]/ul//a/@href')[-1].split('/')[-1]).group()
        return int(page_num)

    def get_chpater_content(self, book_url, page_num):
        # 多请求了4页
        page_dict = {}
        for i in range(1, page_num+5):
            chapter_url = re.sub(r'\d+', str(i), book_url)
            res = self.send_request(chapter_url.replace('https', 'http'), headers=self.headers)
            if res.status_code != 404:
                print(chapter_url)
                html_str = self.bytes_to_str(res.content)
                page_dict[chapter_url] = html_str
        return page_dict

    def get_book_list(self):
        self.struct_book_url()
        print(len(self.book_url))
        for url_info_dict in self.book_url:
            # 获得原始目录的html
            # try:
            star_url = url_info_dict['url_1'].replace('\n', '')
            html_bytes = self.send_request(url=star_url, headers=self.headers_dzkb)
            chpater_html = self.bytes_to_str(html_bytes.content)
            url_info_dict['chapter_html'] = chpater_html

            book_url = re.search(r'https://www.ruiwen.com/(.+)isstop=True', url_info_dict['url_4']).group()
            print('*'*20)
            print(book_url)
            page_num = self.get_new_chapter_list(book_url)
            # 获得每个章节
            page_dict = self.get_chpater_content(book_url, page_num)
            url_info_dict['pages'] = page_dict
            with open(self.current_path_name+'/ruiwen/already.txt', 'a', encoding='utf-8') as f:
                f.write(json.dumps(url_info_dict, ensure_ascii=False)+'\n')
            # except Exception:
            #     with open(self.current_path_name+'/ruiwen/error.txt', 'a', encoding='utf-8') as f:
            #         f.write(json.dumps(url_info_dict, ensure_ascii=False)+'\n')


if __name__ == '__main__':
    ruiwen = BookListRuiWen()
    ruiwen.get_book_list()
