import requests
import re
import os
import json
from retrying import retry
from lxml import etree
from requests import request
from openpyxl import load_workbook


class BookListKeBiao(object):

    def __init__(self):
        self.url_list = []
        self.book_url = []
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36'
        }
        self.headers_dzkb = {
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3',
            'Accept-Encoding': 'gzip, deflate',
            'Accept-Language': 'zh-CN,zh;q=0.9',
            'Connection': 'keep-alive',
            'Host': 'www.dzkbw.com',
            'Upgrade-Insecure-Requests': '1',
            'Cookie': 'ASPSESSIONIDAQSDDTCR=PNODCHKBCOCBGOFPAHEFELJN; _ga=GA1.2.1041419005.1569411852; _gid=GA1.2.1303393327.1569411852; Hm_lvt_463b13a51a26c92e0b49cd9e0dedb52b=1569411854; ASPSESSIONIDASSDCTDQ=LGDBIFLBPJEDOCCFNDDKEPEG; ASPSESSIONIDCSSADBCD=MPAAKLMBIGBFEBCLPNOJJFEM; ASPSESSIONIDCQQCDTCR=BCFFPGACFFKPBEOAABPIIAEN; ASPSESSIONIDCSRBARBS=FCIJEJACMOENGAABODEJJAEC; ASPSESSIONIDCQSBCACD=OAAAOIACBHBMFKPMDNJMOFGP; ASPSESSIONIDCSSADACD=HBGGNJBCGLJFNCKPFJCAHLHP; ASPSESSIONIDCSSACADC=KELCFACCHDIFGLOODAIMBPLP; ASPSESSIONIDASRDCSCR=CGEJILCCAJKCMBCMDMMMKFDD; ASPSESSIONIDASSBDADD=CJBGEDDCOJFGDENNNIBNNIBI; ASPSESSIONIDCSRCDSCQ=IBNOHEDCBCFCCAOEHPIKMGJN; ASPSESSIONIDCQRCBACD=NDOFENDCFJFCJNGMFANONOLD; ASPSESSIONIDCSTAAQBS=NFKAHODCHNABKLECPJKNLMGF; ASPSESSIONIDAQQBAQAT=IDNALKECICKPIABCCHCJOAFA; ASPSESSIONIDASRCDTDR=KPNCEMFCDOJMDKJPNAPKJKHB; ASPSESSIONIDAQSBCBDD=PAGBBBGCIMEHGBBCECHLHFKJ; ASPSESSIONIDASSDBTCR=FLJMCIGCHLJFEDAHJIIMLCOA; ASPSESSIONIDASRBARAT=IKKAJAHCFPEBHGKFKCECFLMH; ASPSESSIONIDCQSCCBCD=HJFFMDHCKPLNJHCCIIDLOLLD; ASPSESSIONIDCSSDCACD=MIDDBCLCPBENHPBOOOAJKMPK; ASPSESSIONIDASSBBQAT=JEDPFJLCIGNPAPJAIKFFNMHC; ASPSESSIONIDAQTDCBCC=LHIIDLLCEGNOJBPPKPPAMECH; ASPSESSIONIDASRCCTCQ=PANMOLLCPLJJGMFCEDANBIDD; ASPSESSIONIDAQTBADDD=ACKKLKMCDEEDFDPKCJIALFMA; ASPSESSIONIDASRDCBDC=NJIKLFNCDLEKDMCNNBHJPFON; ASPSESSIONIDCSTCBBCD=KLMDOMNCNPNAMGKGPPJECOFL; _gat_gtag_UA_66838525_1=1; Hm_lpvt_463b13a51a26c92e0b49cd9e0dedb52b=1569569659',
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.75 Safari/537.36'
        }
        # /Users/guo/Desktop/mywtest/xuezi_net
        self.current_path_name = os.path.dirname(os.path.abspath(__file__))

    def create_file_path(self):
        if not os.path.exists(self.current_path_name+'/kebiao'):
            os.makedirs(self.current_path_name+'/kebiao')
            os.makedirs(self.current_path_name+'/kebiao/pdf')
            os.makedirs(self.current_path_name+'/kebiao/chapter_html')

    @retry(stop_max_attempt_number=6, stop_max_delay=3000, wait_fixed=3000, wait_incrementing_increment=500)
    def send_request(self, url, headers=None):
        requests.packages.urllib3.disable_warnings()
        res = request(method='get', url=url, headers=headers, timeout=20, verify=False)
        return res.content

    def struct_url(self):
        open_excel = load_workbook(os.path.abspath('book_list.xlsx'))
        tab =open_excel['book_list']
        max_num = tab.max_row
        for i in range(1, max_num+1):
            url_1 = tab.cell(i, 1).value
            url_2 = tab.cell(i, 2).value
            url_3 = tab.cell(i, 3).value
            url_4 = tab.cell(i, 4).value
            self.url_list.append({'url_1': url_1, 'url_2': url_2, 'url_3': url_3, 'url_4': url_4})

    # 把字节转成字符串
    def bytes_to_html(self, html_bytes):
        try:
            html_obj = etree.HTML(html_bytes.decode(encoding='utf-8'))
        except Exception:
            try:
                html_obj = etree.HTML(html_bytes.decode(encoding='gbk'))
            except Exception:
                try:
                    html_obj = etree.HTML(html_bytes.decode(encoding='GB2312'))
                except Exception:
                    html_obj = etree.HTML(html_bytes.decode(encoding='GB18030'))
        return html_obj

    # 转换成html字符串
    def bytes_to_str(self, html_bytes):
        try:
            html_obj = html_bytes.decode(encoding='utf-8')
        except Exception:
            try:
                html_obj = html_bytes.decode(encoding='gbk')
            except Exception:
                try:
                    html_obj = html_bytes.decode(encoding='GB2312')
                except Exception:
                    html_obj = html_bytes.decode(encoding='GB18030')
        return html_obj

    def anay_subj(self, html_bytes):
        html_obj = self.bytes_to_html(html_bytes)
        info_list = html_obj.xpath('.//div[@class="daohan"]/*/text()')
        # 获得原始书本的章节目录
        mulu_list = []
        mulu_a_list = html_obj.xpath('.//div[@class="bookmulu"]/a')
        for a_obj in mulu_a_list:
            name = a_obj.xpath('./*/text() | ./text()')[0]
            href = a_obj.xpath('./@href')[0]
            mulu_list.append([name, href])
        return info_list, mulu_list

    # 构造每一本书的请求地址
    def struct_book_url(self):
        # 构造请求的地址
        self.create_file_path()
        self.struct_url()
        for url_info_dict in self.url_list:
            url_str = url_info_dict['url_4']
            url_re_obj = re.search(r'http://www.xscbs.com/.+isstop=True', url_str)
            if url_re_obj:
                self.book_url.append(url_info_dict)

    # 获得书本的每个章节地址
    def get_book(self, url):
        html_bytes = self.send_request(url, self.headers)
        html_obj = self.bytes_to_html(html_bytes)
        a_url_list = []
        a_obj_list = html_obj.xpath('.//div[@class="affff"]/a')
        for a_obj in a_obj_list:
            a_url = 'http://www.xscbs.com' + a_obj.xpath('./@href')[0]
            a_url_list.append({'chapter_url': a_url})
        return a_url_list

    # 获得每个章节的pdf地址, 并下载pdf
    def get_chapter_pdf(self, url):
        html_bytes = self.send_request(url, self.headers)
        html_obj = self.bytes_to_html(html_bytes)
        pdf_url = html_obj.xpath('.//iframe/@src')[0]
        # 下载PDF内容
        pdf_name = pdf_url.split('/')[-1]
        return pdf_name

    # 获得章节和PDF名字对应关系
    def get_chapter_orm(self, chapter_url_list):
        new_chapter_url_list = []
        for chapter_dict in chapter_url_list:
            print(chapter_dict)
            try:
                pdf_name = self.get_chapter_pdf(chapter_dict['chapter_url'])
                chapter_dict['pdf_name'] = pdf_name
                new_chapter_url_list.append(chapter_dict)
            except Exception:
                new_chapter_url_list.append(chapter_dict)
                with open(self.current_path_name + '/kebiao/pdf_name_error.txt', 'a+', encoding='utf-8') as f:
                    f.write(json.dumps(chapter_dict, ensure_ascii=False) + '\n')
        return new_chapter_url_list

    def run(self):
        self.struct_book_url()
        for url_info_dict in self.book_url:
            # 获得原始目录的html
            url_str = url_info_dict['url_4']
            url_re_obj = re.search(r'http://www.xscbs.com/.+isstop=True', url_str)
            try:
                star_url = url_info_dict['url_2'].replace('\n', '')
                html_bytes = self.send_request(url=star_url, headers=self.headers_dzkb)
                html_str = self.bytes_to_str(html_bytes)
                url_info_dict['chapter_html'] = html_str
                page_num = int(re.search('maxPage=(\d+);', html_str).group(1))
                url_info_dict['page_all_num'] = page_num
                # 获得这本书
                book_url = url_re_obj.group()
                chapter_url_list = self.get_book(book_url)
                new_chapter_url_list = self.get_chapter_orm(chapter_url_list)
                url_info_dict['new_chapater'] = new_chapter_url_list
                with open(self.current_path_name+'/kebiao/already.json', 'a+', encoding='utf-8') as f_json:
                    f_json.write(json.dumps(url_info_dict, ensure_ascii=False)+'\n')
            except Exception:
                with open(self.current_path_name+'/kebiao/url_error.txt', 'a+', encoding='utf-8') as f_err:
                    f_err.write(json.dumps(url_info_dict)+'\n')


    def analys_kebiao(self):
        with open(self.current_path_name + '/kebiao/already.json', 'r', encoding='utf-8') as f_json:
            # f_json.write(json.dumps(url_info_dict, ensure_ascii=False) + '\n')
            info_dict_str_list = f_json.readlines()
            for info_dict_str in info_dict_str_list:
                info_dict = eval(info_dict_str)
                print(info_dict['url_2'])
                book_info = {}
                book_info['source_url'] = info_dict['url_2']
                chapter_html = info_dict['chapter_html']
                html_obj = etree.HTML(chapter_html)
                base_info = html_obj.xpath('.//div[@class="daohan"]//a/text()')
                book_info['book_name']='_'.join(base_info)
                if re.search(r'一年级|二年级|三年级|四年级|五年级|六年级', base_info[3]):
                    book_info['period'] = '小学'
                    book_info['grade'] = re.search(r'一年级|二年级|三年级|四年级|五年级|六年级', base_info[3]).group()
                elif re.search(r'七年级|八年级|九年级', base_info[3]):
                    book_info['period'] = '初中'
                    book_info['grade'] = re.search('七年级|八年级|九年级', base_info[3]).group()
                elif re.search(r'高一|高二|高三', base_info[3]):
                    book_info['period'] = '高中'
                    book_info['grade'] = re.search(r'高一|高二|高三', base_info[3]).group()
                book_info['press_version'] = base_info[1]
                book_info['subject'] = base_info[2]
                book_info['catalogs'] = []
                book_info['pages'] = []

                # 提取章节信息
                li_obj_list = html_obj.xpath('.//div[@class="mululist"]//a')
                chapter_list = []
                for li_obj in li_obj_list:
                    name_list = li_obj.xpath('./text()|.//*/text()')
                    if len(name_list) > 1:
                        name = ' '.join(name_list)
                    else:
                        name = name_list[0]
                    chapter_list.append(name)
                # 获取大章节
                big_chapter_list = []
                n = len(chapter_list)
                big_index_list = []
                for chapter_name, i in zip(chapter_list, range(1, n + 1)):
                    if not re.search('^[1-70分加减合叠]|^三位|^一位|^三角|^十个|^三字|^十二', chapter_name):
                        if re.search('封面|目录|上学了|一|二|三|四|五|六|七|八|九|十|附录|单元|汉语拼音|^课文|^识字|生字表', chapter_name[:5]):
                            big_chapter_list.append({'name': chapter_name, 'catalogs': [], 'pages': []})
                            big_index_list.append(i)
                big_index_list.append(n + 1)
                # print(big_chapter_list)

                # 把小章节添加到大章节中
                m = len(big_index_list)
                for pre_big_dcit, j in zip(big_chapter_list, range(0, m)):
                    for little_chaper in chapter_list[big_index_list[j]:big_index_list[j + 1] - 1]:
                        pre_big_dcit['catalogs'].append({'name': little_chaper, 'catalogs': [], 'pages': []})
                # print(json.dumps(big_chapter_list, ensure_ascii=False))

                book_info['catalogs'] = big_chapter_list
                print(json.dumps(book_info, ensure_ascii=False))

                # PDF的名字列表
                # 图片路径 /home/mengyanwei/myw_env/kebiao/pdf_to_img/+pdf名字

                for chapter_dict in info_dict['new_chapater']:
                    pdf_name = chapter_dict['pdf_name'].replace('.pdf', '')
                    img_list = os.listdir('/home/mengyanwei/myw_env/kebiao/pdf_to_img/'+pdf_name)
                    img_list.sort(key=lambda x: int(re.search('_(\d+).jpg', x).group(1)))
                    book_info['catalogs'].extend(img_list)

                print(book_info)
                with open(self.current_path_name + '/kebiao/add_analys_already.json', 'a', encoding='utf-8') as f:
                    f.write(json.dumps(book_info, ensure_ascii=False)+'\n')







if __name__ == '__main__':
    kebiao = BookListKeBiao()
#    kebiao.run()
    kebiao.analys_kebiao()
