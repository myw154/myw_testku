import requests
import re
import os
import json
from retrying import retry
from lxml import etree
from requests import request
from openpyxl import load_workbook
from urllib import parse

class BookListPdf(object):

    def __init__(self):
        self.book_url_list = []
        self.headers = {
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3',
            'Accept-Encoding': 'gzip, deflate',
            'Accept-Language': 'zh-CN,zh;q=0.9',
            'Connection': 'keep-alive',
            'Host': 'www.dzkbw.com',
            'Upgrade-Insecure-Requests': '1',
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.75 Safari/537.36'
        }
        self.current_path_name = os.path.dirname(os.path.abspath(__file__))

    def create_file_path(self):
        if not os.path.exists(self.current_path_name+'/yousi'):
            os.makedirs(self.current_path_name+'/yousi')

    @retry(stop_max_attempt_number=6, stop_max_delay=3000, wait_fixed=3000, wait_incrementing_increment=500)
    def send_request(self, url, headers=None):
        res = request(method='get', url=url, headers=headers, timeout=20)
        return res

    def struct_url(self, re_metch_url):
        open_excel = load_workbook(self.current_path_name+'/book_list.xlsx')
        tab = open_excel['book_list']
        max_num = tab.max_row
        for i in range(1, max_num+1):
            url_1 = tab.cell(i, 1).value
            url_2 = tab.cell(i, 2).value
            url_3 = tab.cell(i, 3).value
            url_4 = tab.cell(i, 4).value
            re_obj = re.search(re_metch_url, url_4)
            if re_obj:
                self.book_url_list.append({'url_1': url_1, 'url_2': url_2, 'url_3': url_3, 'url_4': url_4})

    # 把字节转成字符串
    def bytes_to_html(self, res):
        try:
            html_obj = res.content.decode(encoding='utf-8')
        except Exception:
            try:
                html_obj = res.content.decode(encoding='gbk')
            except Exception:
                try:
                    html_obj = res.content.decode(encoding='GB2312')
                except Exception:
                    html_obj = res.content.decode(encoding='GB18030')
        return html_obj


    def bytes_to_str(self, html_bytes):
        try:
            html_obj = html_bytes.decode(encoding='utf-8')
        except Exception:
            try:
                html_obj = html_bytes.decode(encoding='gbk')
            except Exception:
                try:
                    html_obj = html_bytes.decode(encoding='GB2312')
                except Exception:
                    html_obj = html_bytes.decode(encoding='GB18030')
        return html_obj




    def anay_subj(self, htm_bytes):
        html_obj = etree.HTML(htm_bytes.decode(encoding='GBK'))
        info_list = html_obj.xpath('.//div[@class="daohan"]/*/text()')
        return info_list

    def analy_book_url(self, res):
        html_obj = etree.HTML(res.content.decode(encoding='utf-8'))
        pdf_url_str = parse.quote(html_obj.xpath('.//iframe[@id="ifrPage"]/@src')[0])
        pdf_url_re = re.search('http.+.pdf', pdf_url_str).group()
        pdf_url = parse.unquote(parse.unquote(pdf_url_re))
        return pdf_url

    def run(self):
        self.create_file_path()
        self.struct_url(r'(http://jfpdf.yousi.com/(\d*.pdf))#page|http://jiaofu.yousi.com/lesson.+isstop=True')
        print(len(self.book_url_list))
        for url_info_dict in self.book_url_list:
            book_info_dict = {}
            book_info_dict['url_1'] = url_info_dict['url_1']
            book_info_dict['url_2'] = url_info_dict['url_2']
            book_info_dict['url_3'] = url_info_dict['url_3']
            book_info_dict['url_4'] = url_info_dict['url_4']
            print(url_info_dict['url_4'])
            try:
                if re.search(r'(http://jfpdf.yousi.com/(\d*.pdf))#page', url_info_dict['url_4']):
                    book_pdf_url = re.search(r'(http://jfpdf.yousi.com/(\d*.pdf))#page', url_info_dict['url_4'])
                    # book科目解析
                    res = self.send_request(url=url_info_dict['url_1'].replace('\n', ''), headers=self.headers)
                    book_info_dict['chapter_html'] = self.bytes_to_str(res.content)
                    book_info_dict['pdf_url'] = book_pdf_url.group()
                    with open(os.path.join(self.current_path_name, 'yousi/new_already.json'), 'a', encoding='utf-8') as f:
                        f.write(json.dumps(book_info_dict, ensure_ascii=False)+'\n')
                else:
                    book_url = re.search(r'http://jiaofu.yousi.com/lesson.+isstop=True', url_info_dict['url_4'])
                    # book科目解析
                    res = self.send_request(url=url_info_dict['url_1'].replace('\n', ''), headers=self.headers)
                    html_bytes = self.send_request(url=book_url.group())
                    pdf_url = self.analy_book_url(html_bytes)
                    book_info_dict['chapter_html'] = self.bytes_to_str(res.content) 
                    book_info_dict['pdf_url'] = pdf_url
                    with open(os.path.join(self.current_path_name, 'yousi/new_already.json'), 'a', encoding='utf-8') as f:
                        f.write(json.dumps(book_info_dict, ensure_ascii=False)+'\n')
            except Exception:
                with open(os.path.join(self.current_path_name, 'yousi/url_error.txt'), 'a+', encoding='utf-8') as f:
                    f.write(str(url_info_dict) + '\n')

if __name__ == '__main__':
    booklist_pdf = BookListPdf()
    booklist_pdf.run()
