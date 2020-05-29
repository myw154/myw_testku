import requests
import re
import os
import json
from retrying import retry
from lxml import etree
from requests import request
from openpyxl import load_workbook


class BookListKeJian(object):

    def __init__(self):
        self.url_list = []
        self.book_url = []
        self.book_first_url = []
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36'
        }
        self.headers_dzkb = {
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3',
            'Accept-Encoding': 'gzip, deflate',
            'Accept-Language': 'zh-CN,zh;q=0.9',
            'Connection': 'keep-alive',
            'Host': 'www.dzkbw.com',
            'Upgrade-Insecure-Requests': '1',
            'Cookie': 'ASPSESSIONIDAQSDDTCR=PNODCHKBCOCBGOFPAHEFELJN; _ga=GA1.2.1041419005.1569411852; _gid=GA1.2.1303393327.1569411852; Hm_lvt_463b13a51a26c92e0b49cd9e0dedb52b=1569411854; ASPSESSIONIDASSDCTDQ=LGDBIFLBPJEDOCCFNDDKEPEG; ASPSESSIONIDCSSADBCD=MPAAKLMBIGBFEBCLPNOJJFEM; ASPSESSIONIDCQQCDTCR=BCFFPGACFFKPBEOAABPIIAEN; ASPSESSIONIDCSRBARBS=FCIJEJACMOENGAABODEJJAEC; ASPSESSIONIDCQSBCACD=OAAAOIACBHBMFKPMDNJMOFGP; ASPSESSIONIDCSSADACD=HBGGNJBCGLJFNCKPFJCAHLHP; ASPSESSIONIDCSSACADC=KELCFACCHDIFGLOODAIMBPLP; ASPSESSIONIDASRDCSCR=CGEJILCCAJKCMBCMDMMMKFDD; ASPSESSIONIDASSBDADD=CJBGEDDCOJFGDENNNIBNNIBI; ASPSESSIONIDCSRCDSCQ=IBNOHEDCBCFCCAOEHPIKMGJN; ASPSESSIONIDCQRCBACD=NDOFENDCFJFCJNGMFANONOLD; ASPSESSIONIDCSTAAQBS=NFKAHODCHNABKLECPJKNLMGF; ASPSESSIONIDAQQBAQAT=IDNALKECICKPIABCCHCJOAFA; ASPSESSIONIDASRCDTDR=KPNCEMFCDOJMDKJPNAPKJKHB; ASPSESSIONIDAQSBCBDD=PAGBBBGCIMEHGBBCECHLHFKJ; ASPSESSIONIDASSDBTCR=FLJMCIGCHLJFEDAHJIIMLCOA; ASPSESSIONIDASRBARAT=IKKAJAHCFPEBHGKFKCECFLMH; ASPSESSIONIDCQSCCBCD=HJFFMDHCKPLNJHCCIIDLOLLD; ASPSESSIONIDCSSDCACD=MIDDBCLCPBENHPBOOOAJKMPK; ASPSESSIONIDASSBBQAT=JEDPFJLCIGNPAPJAIKFFNMHC; ASPSESSIONIDAQTDCBCC=LHIIDLLCEGNOJBPPKPPAMECH; ASPSESSIONIDASRCCTCQ=PANMOLLCPLJJGMFCEDANBIDD; ASPSESSIONIDAQTBADDD=ACKKLKMCDEEDFDPKCJIALFMA; ASPSESSIONIDASRDCBDC=NJIKLFNCDLEKDMCNNBHJPFON; ASPSESSIONIDCSTCBBCD=KLMDOMNCNPNAMGKGPPJECOFL; _gat_gtag_UA_66838525_1=1; Hm_lpvt_463b13a51a26c92e0b49cd9e0dedb52b=1569569659',
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.75 Safari/537.36'
        }
        # /Users/guo/Desktop/mywtest/xuezi_net
        self.current_path_name = os.path.dirname(os.path.abspath(__file__))

    def create_file_path(self):
        if not os.path.exists(self.current_path_name+'/kejian'):
            os.makedirs(self.current_path_name+'/kejian')
        if not os.path.exists(self.current_path_name+'/kejian/img'):
            os.makedirs(self.current_path_name+'/kejian/img')
        if not os.path.exists(self.current_path_name+'/kejian/chapter_html'):
            os.makedirs(self.current_path_name+'/kejian/chapter_html')

    @retry(stop_max_attempt_number=6, stop_max_delay=3000, wait_fixed=3000, wait_incrementing_increment=500)
    def send_request(self, url, headers=None):
        requests.packages.urllib3.disable_warnings()
        res = request(method='get', url=url, headers=headers, timeout=20, verify=False)
        return res.content

    def struct_url(self):
        open_excel = load_workbook(os.path.abspath('book_list.xlsx'))
        tab =open_excel['book_list']
        max_num = tab.max_row
        # for i in range():
        for i in range(1, max_num+1):
            url_1 = tab.cell(i, 1).value
            url_2 = tab.cell(i, 2).value
            url_3 = tab.cell(i, 3).value
            url_4 = tab.cell(i, 4).value
            self.url_list.append({'url_1': url_1, 'url_2': url_2, 'url_3': url_3, 'url_4': url_4})

    # 把字节转成hrml对象
    def bytes_to_html(self, html_bytes):
        try:
            html_obj = etree.HTML(html_bytes.decode(encoding='utf-8'))
        except Exception:
            try:
                html_obj = etree.HTML(html_bytes.decode(encoding='gbk'))
            except Exception:
                try:
                    html_obj = etree.HTML(html_bytes.decode(encoding='GB2312'))
                except Exception:
                    html_obj = etree.HTML(html_bytes.decode(encoding='GB18030'))
        return html_obj

    # 转换成html字符串
    def bytes_to_str(self, html_bytes):
        try:
            html_obj = html_bytes.decode(encoding='utf-8')
        except Exception:
            try:
                html_obj = html_bytes.decode(encoding='gbk')
            except Exception:
                try:
                    html_obj = html_bytes.decode(encoding='GB2312')
                except Exception:
                    html_obj = html_bytes.decode(encoding='GB18030')
        return html_obj

    def anay_subj(self, html_bytes):
        html_obj = self.bytes_to_html(html_bytes)
        info_list = html_obj.xpath('.//div[@class="daohan"]/*/text()')
        # 获得原始书本的章节目录
        mulu_list = []
        mulu_a_list = html_obj.xpath('.//div[@class="bookmulu"]/a')
        for a_obj in mulu_a_list:
            name = a_obj.xpath('./*/text() | ./text()')[0] if len(a_obj.xpath('./*/text() | ./text()')) > 0 else None
            href = 'http://www.dzkbw.com'+a_obj.xpath('./@href')[0]
            mulu_list.append([name, href])
        return info_list, mulu_list

    # 构造每一本书的请求地址
    def struct_book_url(self):
        # 构造请求的地址
        self.create_file_path()
        self.struct_url()
        for url_info_dict in self.url_list:
            url_str = url_info_dict['url_4']
            url_re_obj = re.search('http://www\.kjzhan\.com/(.+)\.html\?isstop=True', url_str)
            if url_re_obj:
                self.book_url.append(url_info_dict)

    # 获得书本的页码总数
    def get_book(self, url):
        res_bytes = self.send_request(url, self.headers)
        html_str = res_bytes.decode(encoding='GBK')
        page_num = int(re.search('totalpage = (\d+);', html_str).group(1))
        return page_num

    # 获得书本的每一页书本图片地址信息
    def get_detail_page(self, url, page_num):
        pages_dict = {}
        for i in range(1, page_num+1):
            if i == 1:
                detail_url = re.sub('_(\d+)\.html', '.html', url)
            else:
                detail_url = re.sub('_(\d+)\.html', '_'+str(i)+'.html', url)
            res_bytes = self.send_request(detail_url, self.headers)
            html_obj = self.bytes_to_str(res_bytes)
            pages_dict[detail_url] = html_obj
        return pages_dict


    # 获得每本书的目录信息，获得每本书的第一页的url地址
    def get_book_list(self):
        self.struct_book_url()
        print(len(self.book_url))
        for url_info_dict in self.book_url:
            # 获得原始目录的html
            try:
                # 提取基本信息用
                star_url = url_info_dict['url_1'].replace('\n', '')
                html_bytes = self.send_request(url=star_url, headers=self.headers_dzkb)
                html_str = self.bytes_to_str(html_bytes)
                url_info_dict['chapter_html'] = html_str
                # 提取章节用
                url = url_info_dict['url_2'].replace('\n', '')
                html_bytes_2 = self.send_request(url=url, headers=self.headers_dzkb)
                html_str_2 = self.bytes_to_str(html_bytes_2)
                url_info_dict['chapter_html_2'] = html_str_2

                book_url = re.search(r'http://www.kjzhan.com/(.+).html\?isstop=True', url_info_dict['url_4']).group()
                print(book_url)
                # 获得总页数
                page_num = self.get_book(book_url)
                pages_dict = self.get_detail_page(book_url, page_num)
                url_info_dict['pages_all_num'] = page_num
                url_info_dict['pages'] = pages_dict
                with open(self.current_path_name+'/kejian/already_get_book.txt', 'a+', encoding='utf-8') as f:
                    f.write(json.dumps(url_info_dict, ensure_ascii=False)+'\n')
            except Exception:
                with open(self.current_path_name+'/kejian/error_url.txt', 'a+', encoding='utf-8') as f:
                    f.write(json.dumps(url_info_dict)+'\n')


if __name__ == '__main__':
    kejian = BookListKeJian()
    kejian.get_book_list()