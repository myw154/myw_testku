import requests
import re
import os
import json
from retrying import retry
from lxml import etree
from requests import request
from openpyxl import load_workbook


class BookList10087(object):

    def __init__(self):
        self.url_list = []
        self.book_url = []
        self.book_first_url = []
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36'
        }
        self.headers_dzkb = {
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3',
            'Accept-Encoding': 'gzip, deflate',
            'Accept-Language': 'zh-CN,zh;q=0.9',
            'Connection': 'keep-alive',
            'Host': 'www.dzkbw.com',
            'Upgrade-Insecure-Requests': '1',
            'Cookie': 'ASPSESSIONIDAQSDDTCR=PNODCHKBCOCBGOFPAHEFELJN; _ga=GA1.2.1041419005.1569411852; _gid=GA1.2.1303393327.1569411852; Hm_lvt_463b13a51a26c92e0b49cd9e0dedb52b=1569411854; ASPSESSIONIDASSDCTDQ=LGDBIFLBPJEDOCCFNDDKEPEG; ASPSESSIONIDCSSADBCD=MPAAKLMBIGBFEBCLPNOJJFEM; ASPSESSIONIDCQQCDTCR=BCFFPGACFFKPBEOAABPIIAEN; ASPSESSIONIDCSRBARBS=FCIJEJACMOENGAABODEJJAEC; ASPSESSIONIDCQSBCACD=OAAAOIACBHBMFKPMDNJMOFGP; ASPSESSIONIDCSSADACD=HBGGNJBCGLJFNCKPFJCAHLHP; ASPSESSIONIDCSSACADC=KELCFACCHDIFGLOODAIMBPLP; ASPSESSIONIDASRDCSCR=CGEJILCCAJKCMBCMDMMMKFDD; ASPSESSIONIDASSBDADD=CJBGEDDCOJFGDENNNIBNNIBI; ASPSESSIONIDCSRCDSCQ=IBNOHEDCBCFCCAOEHPIKMGJN; ASPSESSIONIDCQRCBACD=NDOFENDCFJFCJNGMFANONOLD; ASPSESSIONIDCSTAAQBS=NFKAHODCHNABKLECPJKNLMGF; ASPSESSIONIDAQQBAQAT=IDNALKECICKPIABCCHCJOAFA; ASPSESSIONIDASRCDTDR=KPNCEMFCDOJMDKJPNAPKJKHB; ASPSESSIONIDAQSBCBDD=PAGBBBGCIMEHGBBCECHLHFKJ; ASPSESSIONIDASSDBTCR=FLJMCIGCHLJFEDAHJIIMLCOA; ASPSESSIONIDASRBARAT=IKKAJAHCFPEBHGKFKCECFLMH; ASPSESSIONIDCQSCCBCD=HJFFMDHCKPLNJHCCIIDLOLLD; ASPSESSIONIDCSSDCACD=MIDDBCLCPBENHPBOOOAJKMPK; ASPSESSIONIDASSBBQAT=JEDPFJLCIGNPAPJAIKFFNMHC; ASPSESSIONIDAQTDCBCC=LHIIDLLCEGNOJBPPKPPAMECH; ASPSESSIONIDASRCCTCQ=PANMOLLCPLJJGMFCEDANBIDD; ASPSESSIONIDAQTBADDD=ACKKLKMCDEEDFDPKCJIALFMA; ASPSESSIONIDASRDCBDC=NJIKLFNCDLEKDMCNNBHJPFON; ASPSESSIONIDCSTCBBCD=KLMDOMNCNPNAMGKGPPJECOFL; _gat_gtag_UA_66838525_1=1; Hm_lpvt_463b13a51a26c92e0b49cd9e0dedb52b=1569569659',
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.75 Safari/537.36'
        }
        # /Users/guo/Desktop/mywtest/xuezi_net
        self.current_path_name = os.path.dirname(os.path.abspath(__file__))

    def create_file_path(self):
        if not os.path.exists(self.current_path_name+'/basic_net'):
            os.makedirs(self.current_path_name+'/basic_net')
        if not os.path.exists(self.current_path_name+'/basic_net/img'):
            os.makedirs(self.current_path_name+'/basic_net/img')
        if not os.path.exists(self.current_path_name+'/basic_net/book_chapter_html'):
            os.makedirs(self.current_path_name+'/basic_net/book_chapter_html')
        if not os.path.exists(self.current_path_name + '/basic_net/original_chapter_html/'):
            os.makedirs(self.current_path_name + '/basic_net/original_chapter_html/')

    @retry(stop_max_attempt_number=5, stop_max_delay=3000, wait_fixed=3000, wait_incrementing_increment=500)
    def send_request(self, url, headers=None):
        requests.packages.urllib3.disable_warnings()
        res = request(method='get', url=url, headers=headers, timeout=10, verify=False)
        return res

    @retry(stop_max_attempt_number=5, stop_max_delay=3000, wait_fixed=3000, wait_incrementing_increment=500)
    def send_request_post(self, url, data, headers=None):
        requests.packages.urllib3.disable_warnings()
        res = request(method='post', url=url, data=data, headers=headers, timeout=10, verify=False)
        return res

    def struct_url(self):
        open_excel = load_workbook(os.path.abspath('book_list.xlsx'))
        tab = open_excel['book_list']
        max_num = tab.max_row
        # for i in range(888, 889):
        for i in range(1, max_num+1):
            url_1 = tab.cell(i, 1).value
            url_2 = tab.cell(i, 2).value
            url_3 = tab.cell(i, 3).value
            url_4 = tab.cell(i, 4).value
            re_obj = re.search(r'http://www.100875.com.cn/(.*)', url_4)
            if re_obj:
                self.book_url.append({'url_1': url_1, 'url_2': url_2, 'url_3': url_3, 'url_4': url_4})

    # 把字节转成字符串
    def bytes_to_html(self, html_bytes):
        try:
            html_obj = html_bytes.content.decode(encoding='utf-8')
        except Exception:
            try:
                html_obj = html_bytes.content.decode(encoding='gbk')
            except Exception:
                try:
                    html_obj = html_bytes.content.decode(encoding='GB2312')
                except Exception:
                    html_obj = html_bytes.content.decode(encoding='GB18030')
        return html_obj

    # 转换成html字符串
    def bytes_to_str(self, html_bytes):
        try:
            html_obj = html_bytes.decode(encoding='utf-8')
        except Exception:
            try:
                html_obj = html_bytes.decode(encoding='gbk')
            except Exception:
                try:
                    html_obj = html_bytes.decode(encoding='GB2312')
                except Exception:
                    html_obj = html_bytes.decode(encoding='GB18030')
        return html_obj

    def anay_subj(self, html_bytes):
        html_obj = self.bytes_to_html(html_bytes)
        info_list = html_obj.xpath('.//div[@class="daohan"]/*/text()')
        # 获得原始书本的章节目录
        mulu_list = []
        mulu_a_list = html_obj.xpath('.//div[@class="bookmulu"]/a')
        for a_obj in mulu_a_list:
            name = a_obj.xpath('./*/text() | ./text()')[0] if len(a_obj.xpath('./*/text() | ./text()')) > 0 else None
            href = 'http://www.dzkbw.com'+a_obj.xpath('./@href')[0]
            mulu_list.append([name, href])
        return info_list, mulu_list

    # 构造每一本书的请求地址
    def struct_book_url(self):
        # 创建一些必要的文件夹路径
        self.create_file_path()
        # 构造书本的地址
        self.struct_url()

    # 获得新的章节目录
    def get_new_mulu_list(self, book_url):
        data_dict = dict()
        resId = re.search(r'resId=(.+?)&', book_url).group(1)
        data_dict['resId'] = resId
        res_byetes = request(method='post', url='http://www.100875.com.cn/ebook/catalogList', data=data_dict, headers=self.headers)
        res_str = res_byetes.content.decode(encoding='utf-8')
        chapter_info_dict = json.loads(res_str)['data']
        return chapter_info_dict

    # 获得所有章节信息
    def get_all_chapter(self, chapter_url_list, contributeId):
        chapter_url_list_new = []
        for chapter_url_dict in chapter_url_list:
                info_list = self.get_chapter(chapter_url_dict, contributeId)
                chapter_url_dict['chapter_info'] = info_list
                chapter_url_list_new.append(chapter_url_dict)
        return chapter_url_list_new

    # 获得一个章节信息
    def get_chapter(self, chapter_url_dict, contributeId):
        data = dict()
        data['bookid'] = chapter_url_dict['bookId']
        data['contributeId'] = contributeId
        # res_bytes = request(method='post', url='http://www.100875.com.cn/ebook/bookPictureList', headers=self.headers, data=data)
        res_bytes = self.send_request_post(url='http://www.100875.com.cn/ebook/bookPictureList',data=data)
        info_list = json.loads(res_bytes.content.decode(encoding='utf-8'))['data']
        return info_list

    def get_book_list_old(self):
        self.struct_book_url()
        print(len(self.book_url))
        for url_info_dict in self.book_url:
            # 获得原始目录的html
            try:
                star_url = url_info_dict['url_1'].replace('\n', '')
                res = self.send_request(url=star_url, headers=self.headers_dzkb)
                url_info_dict['chapter_html'] = self.bytes_to_str(res.content)
                # 获取书本的目录
                book_url = re.search(r'http://www.100875.com.cn/(.+)&isstop=True', url_info_dict['url_4']).group().replace('&isstop=True', '')
                chapter_info_dict_list = self.get_new_mulu_list(book_url)
                contributeId = re.search(r'contributeId=(\d+)', book_url).group(1)
                print(book_url)
                for chapter_info_dic in chapter_info_dict_list:
                    chapter_list = chapter_info_dic['catalogList']
                    self.get_all_chapter(chapter_list, contributeId)
                with open(self.current_path_name+'/basic_net/already.txt', 'a', encoding='utf-8') as f:
                    f.write(json.dumps(url_info_dict, ensure_ascii=False)+'\n')
            except Exception:
                with open(self.current_path_name+'/basic_net/error.txt', 'a', encoding='utf-8') as f:
                    f.write(json.dumps(url_info_dict, ensure_ascii=False)+'\n')

    def get_book_list(self):
        self.struct_book_url()
        print(len(self.book_url))
        for url_info_dict in self.book_url:
            # 获得原始目录的html
            try:
                star_url = url_info_dict['url_1'].replace('\n', '')
                res_obj = self.send_request(url=star_url, headers=self.headers_dzkb)
                url_info_dict['chpater_html'] = self.bytes_to_str(res_obj.content)
                # 获得书本的章节目录
                book_url = re.search(r'http://www.100875.com.cn/(.+)&isstop=True', url_info_dict['url_4']).group().replace('&isstop=True', '')
                chapter_info_dict_list = self.get_new_mulu_list(book_url)
                contributeId = re.search(r'contributeId=(\d+)', book_url).group(1)
                print(book_url)
                chapter_info_list = []
                for chapter_info_dic in chapter_info_dict_list:
                    chapter_list = chapter_info_dic['catalogList']
                    chapeter_all_list = self.get_all_chapter(chapter_list, contributeId)
                    chapter_info_list.append(chapeter_all_list)
                url_info_dict['chapter'] = chapter_info_list
                with open(self.current_path_name+'/basic_net/already.txt', 'a', encoding='utf-8') as f:
                    f.write(json.dumps(url_info_dict, ensure_ascii=False)+'\n')
            except Exception:
                with open(self.current_path_name+'/basic_net/error.txt', 'a', encoding='utf-8') as f:
                    f.write(json.dumps(url_info_dict, ensure_ascii=False)+'\n')


if __name__ == '__main__':
    basic_100875 = BookList10087()
    basic_100875.get_book_list()
